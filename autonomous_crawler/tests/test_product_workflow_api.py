from __future__ import annotations

import json
import tempfile
import time
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

from fastapi.testclient import TestClient

from autonomous_crawler.api.app import create_app, _clear_jobs
from autonomous_crawler.models.product import ProductRecord
from autonomous_crawler.runtime import RuntimeRequest, RuntimeResponse
from autonomous_crawler.runners import ProfileLongRunConfig, run_profile_longrun
from autonomous_crawler.runners.managed_state import build_managed_crawl_state, compact_managed_state_for_llm
from autonomous_crawler.runners.product_workflow import (
    ExportSpec,
    ExportTemplate,
    analyze_site_for_product_workflow,
    _build_progress_summary,
    _derive_current_stage,
    _derive_quality_indicator,
    _extract_last_error_snippet,
    _looks_like_catalog_url,
    _profile_runtime_mode,
    build_run_spec,
    discover_catalog_from_site_fallbacks,
    export_product_records,
    import_catalog_tree,
    profile_from_run_spec,
    resolve_fields,
    summarize_run_progress,
)
from autonomous_crawler.tools.fetch_policy import BestFetchResult, FetchAttempt
from autonomous_crawler.storage.product_store import ProductStore


EXTRACTOR_FIXTURE_ROOT = (
    Path(__file__).resolve().parents[2]
    / "dev_logs"
    / "training"
    / "xiaomi_recon_2026_05_28"
    / "fixtures"
)


def _read_extractor_fixture_text(*parts: str) -> str:
    return EXTRACTOR_FIXTURE_ROOT.joinpath(*parts).read_text(encoding="utf-8")


def _read_extractor_fixture_json(*parts: str):
    return json.loads(_read_extractor_fixture_text(*parts))


class HtmlFixtureFetchRuntime:
    name = "html_fixture"

    def __init__(self, pages: dict[str, str]) -> None:
        self.pages = pages
        self.requests: list[RuntimeRequest] = []

    def fetch(self, request: RuntimeRequest) -> RuntimeResponse:
        self.requests.append(request)
        html = self.pages.get(request.url, "")
        return RuntimeResponse(
            ok=bool(html),
            final_url=request.url,
            status_code=200 if html else 404,
            html=html,
            text=html,
            engine_result={"engine": self.name},
            error="" if html else "not found",
        )


class ProductWorkflowCoreTests(unittest.TestCase):
    def test_import_catalog_tree_accepts_spider_nested_menu_shape(self) -> None:
        payload = {
            "Kobieta": {
                "Produkty": {
                    "Legginsy": "https://shop.test/leggings",
                    "Bluzy": "https://shop.test/hoodies",
                }
            }
        }

        tree = import_catalog_tree(payload)

        leaf = tree[0]["children"][0]["children"][0]
        self.assertEqual(leaf["label"], "Legginsy")
        self.assertEqual(leaf["url"], "https://shop.test/leggings")
        self.assertEqual(leaf["level1"], "Kobieta")
        self.assertEqual(leaf["level2"], "Produkty")
        self.assertEqual(leaf["level3"], "Legginsy")

    def test_resolve_fields_maps_chinese_natural_language_to_canonical_fields(self) -> None:
        available = [
            {"name": "title", "label": "商品标题"},
            {"name": "highest_price", "label": "最高价格"},
            {"name": "colors", "label": "颜色"},
        ]

        result = resolve_fields(available, natural_language="我要标题、原价和颜色")

        self.assertEqual(result["selected_fields"], ["colors", "highest_price", "title"])
        self.assertFalse(result["needs_refinement"])

    @patch("autonomous_crawler.runners.product_workflow._fetch_text_quietly")
    def test_catalog_fallback_extracts_js_route_candidates(self, mock_fetch) -> None:
        html = '<html><script src="/client.js"></script></html>'
        mock_fetch.side_effect = lambda url, max_chars: (
            'const routes=["/rowery","/akcesoria","/sign-in","/cart"];'
            if str(url).endswith("/client.js")
            else ""
        )

        tree = discover_catalog_from_site_fallbacks("https://shop.test/", html)
        urls = {node["url"] for node in tree}

        self.assertIn("https://shop.test/rowery", urls)
        self.assertIn("https://shop.test/akcesoria", urls)
        self.assertNotIn("https://shop.test/cart", urls)

    def test_catalog_url_match_uses_tokens_not_short_substrings(self) -> None:
        self.assertFalse(_looks_like_catalog_url("https://romet.pl/payment-information", "payment information"))
        self.assertFalse(_looks_like_catalog_url("https://romet.pl/saved-payments", "saved payments"))
        self.assertFalse(_looks_like_catalog_url("https://romet.pl/payment-methods", "payment methods"))
        self.assertFalse(_looks_like_catalog_url("https://romet.pl/bike-service", "bike service"))
        self.assertTrue(_looks_like_catalog_url("https://romet.pl/rowery", "rowery"))
        self.assertTrue(_looks_like_catalog_url("https://romet.pl/akcesoria", "akcesoria"))

    @patch("autonomous_crawler.runners.product_workflow.httpx.Client")
    def test_catalog_fallback_prefers_magento_graphql_menu(self, mock_client_cls) -> None:
        response = mock_client_cls.return_value.__enter__.return_value.post.return_value
        response.status_code = 200
        response.json.return_value = {
            "data": {
                "storeConfig": {
                    "secure_base_url": "https://romet.pl/",
                    "category_url_suffix": ".html",
                },
                "categoryList": [
                    {
                        "name": "Root",
                        "include_in_menu": True,
                        "url_path": "",
                        "children": [
                            {
                                "name": "Rowery",
                                "include_in_menu": True,
                                "url_path": "rowery",
                                "children": [
                                    {
                                        "name": "Gorskie",
                                        "include_in_menu": True,
                                        "url_path": "rowery/gorskie",
                                        "children": [],
                                    }
                                ],
                            },
                            {
                                "name": "Payment Information",
                                "include_in_menu": True,
                                "url_path": "payment-information",
                                "children": [],
                            },
                        ],
                    }
                ],
            }
        }

        tree = discover_catalog_from_site_fallbacks("https://romet.pl/", "")
        urls = {node["url"] for node in tree}
        child_urls = {child["url"] for node in tree for child in node.get("children", [])}

        self.assertIn("https://romet.pl/rowery.html", urls)
        self.assertIn("https://romet.pl/rowery/gorskie.html", child_urls)
        self.assertNotIn("https://romet.pl/payment-information.html", urls)

    def test_run_spec_detail_urls_are_executed_as_detail_pages(self) -> None:
        spec = build_run_spec({
            "target_url": "https://shop.test/",
            "catalog_nodes": [
                {
                    "label": "Alpha Product",
                    "url": "https://shop.test/product/alpha-123.html",
                    "path": ["Alpha Product"],
                }
            ],
            "selected_fields": ["title", "highest_price", "image_urls"],
        })

        profile = profile_from_run_spec(spec)
        data = profile.to_dict()

        self.assertEqual(data["crawl_preferences"]["seed_kind"], "detail")
        self.assertEqual(data["crawl_preferences"]["seed_urls"], ["https://shop.test/product/alpha-123.html"])
        self.assertIn("title", data["selectors"]["detail"])
        self.assertIn("image_urls", data["selectors"]["detail"])

    def test_run_spec_list_urls_get_link_discovery_and_detail_fallbacks(self) -> None:
        spec = build_run_spec({
            "target_url": "https://shop.test/",
            "catalog_nodes": [
                {
                    "label": "Shoes",
                    "url": "https://shop.test/collections/shoes",
                    "path": ["Shoes"],
                }
            ],
            "selected_fields": ["title", "description"],
        })

        profile = profile_from_run_spec(spec)
        data = profile.to_dict()

        self.assertEqual(data["crawl_preferences"]["seed_kind"], "list")
        self.assertEqual(data["pagination_hints"]["type"], "dom_links")
        self.assertIn("link_discovery", data["pagination_hints"])
        self.assertIn("title", data["selectors"]["detail"])
        self.assertTrue(any("list seeds require link discovery" in note for note in data["training_notes"]))

    def test_generated_profile_can_crawl_imported_list_to_product_record(self) -> None:
        spec = build_run_spec({
            "target_url": "https://shop.test/",
            "catalog_nodes": [
                {"label": "Shoes", "url": "https://shop.test/collections/shoes", "path": ["Shoes"]}
            ],
            "selected_fields": ["title", "description", "image_urls"],
        })
        profile = profile_from_run_spec(spec)
        fetch = HtmlFixtureFetchRuntime({
            "https://shop.test/collections/shoes": """
                <html><body>
                  <a href="/product/alpha-123.html">Alpha Shoe</a>
                </body></html>
            """,
            "https://shop.test/product/alpha-123.html": """
                <html><head>
                  <meta property="og:title" content="Alpha Shoe">
                  <meta property="og:image" content="https://shop.test/alpha.jpg">
                  <meta name="description" content="A durable running shoe">
                </head><body><h1>Alpha Shoe</h1></body></html>
            """,
        })

        result = run_profile_longrun(
            profile=profile,
            config=ProfileLongRunConfig(run_id="imported-list-smoke", batch_size=10, max_batches=2),
            fetch_runtime=fetch,
        )

        self.assertEqual(result.product_stats["total"], 1)
        self.assertEqual(result.sample_records[0]["title"], "Alpha Shoe")
        self.assertEqual(
            [request.url for request in fetch.requests],
            ["https://shop.test/collections/shoes", "https://shop.test/product/alpha-123.html"],
        )

    def test_generated_profile_can_crawl_imported_detail_url_directly(self) -> None:
        spec = build_run_spec({
            "target_url": "https://shop.test/",
            "catalog_nodes": [
                {"label": "Alpha", "url": "https://shop.test/product/alpha-123.html", "path": ["Alpha"]}
            ],
            "selected_fields": ["title", "description"],
        })
        profile = profile_from_run_spec(spec)
        fetch = HtmlFixtureFetchRuntime({
            "https://shop.test/product/alpha-123.html": """
                <html><head><meta property="og:title" content="Alpha Shoe"></head>
                <body><p class="description">A durable running shoe</p></body></html>
            """,
        })

        result = run_profile_longrun(
            profile=profile,
            config=ProfileLongRunConfig(run_id="imported-detail-smoke", batch_size=10, max_batches=1),
            fetch_runtime=fetch,
        )

        self.assertEqual(result.product_stats["total"], 1)
        self.assertEqual(fetch.requests[0].url, "https://shop.test/product/alpha-123.html")
        self.assertEqual(fetch.requests[0].selectors[0].name, "title")

    def test_profile_from_run_spec_promotes_js_shell_runs_to_dynamic_mode(self) -> None:
        spec = build_run_spec({
            "target_url": "https://romet.pl/",
            "profile": {
                "name": "romet.pl",
                "training_notes": ["PWA shell detected"],
                "crawl_preferences": {"seed_urls": ["https://romet.pl/rowery/rowery-gorskie/mtb.html"]},
            },
            "catalog_nodes": [
                {"label": "MTB", "url": "https://romet.pl/rowery/rowery-gorskie/mtb.html", "path": ["Rowery", "MTB"]}
            ],
            "selected_fields": ["title", "highest_price"],
        })
        profile = profile_from_run_spec(spec)
        data = profile.to_dict()

        self.assertEqual(_profile_runtime_mode(profile), "dynamic")
        self.assertEqual(data["access_config"]["mode"], "dynamic")
        self.assertTrue(data["access_config"]["browser_config"]["auto_accept_cookies"])
        self.assertEqual(data["access_config"]["browser_config"]["render_time_ms"], 5000)

    def test_imported_catalog_can_be_enriched_with_graphql_uid(self) -> None:
        imported = [{
            "id": "user-mtb",
            "label": "MTB",
            "url": "https://romet.pl/rowery/rowery-gorskie/mtb.html",
            "path": ["Rowery", "Rowery Górskie", "MTB"],
        }]
        discovered = [{
            "id": "agent-mtb",
            "label": "MTB",
            "url": "https://romet.pl/rowery/rowery-gorskie/mtb.html",
            "path": ["Rowery", "Rowery Górskie", "MTB"],
            "source": "graphql",
            "graphql_category_uid": "NQ==",
            "children": [],
        }]

        from autonomous_crawler.runners.product_workflow import _enrich_catalog_with_discovered_metadata

        enriched = _enrich_catalog_with_discovered_metadata(imported, discovered)
        spec = build_run_spec({
            "target_url": "https://romet.pl/",
            "profile": {"name": "romet.pl"},
            "catalog_nodes": enriched,
            "selected_fields": ["title", "highest_price"],
            "test_limit": 20,
        })
        profile = profile_from_run_spec(spec, limit=20)

        self.assertEqual(enriched[0]["graphql_category_uid"], "NQ==")
        self.assertEqual(profile.api_hints["endpoint"], "https://romet.pl/graphql")
        self.assertEqual(profile.crawl_preferences["seed_kind"], "api")

    def test_profile_from_run_spec_recovers_graphql_uid_for_stale_plain_catalog(self) -> None:
        plain_catalog = [{
            "id": "stale-llm-mtb",
            "label": "MTB",
            "url": "https://romet.pl/rowery/rowery-gorskie/mtb.html",
            "path": ["Rowery", "Rowery górskie", "MTB"],
            "source": "llm",
        }]
        spec = build_run_spec({
            "target_url": "https://romet.pl/",
            "profile": {
                "name": "romet.pl",
                "crawl_preferences": {
                    "catalog_tree": plain_catalog,
                    "seed_urls": ["https://romet.pl/rowery/rowery-gorskie/mtb.html"],
                    "seed_kind": "list",
                },
            },
            "catalog_nodes": plain_catalog,
            "selected_fields": ["title", "highest_price"],
            "test_limit": 20,
        })

        profile = profile_from_run_spec(spec, limit=20)

        self.assertEqual(profile.api_hints["endpoint"], "https://romet.pl/graphql")
        self.assertEqual(profile.crawl_preferences["seed_kind"], "api")
        self.assertEqual(profile.crawl_preferences["seed_urls"], ["https://romet.pl/graphql"])

    def test_site_analysis_discovers_executable_extraction_contract_from_html(self) -> None:
        html = _read_extractor_fixture_text("superdry_com", "raw_evidence_list_page.html")
        fetch = BestFetchResult(
            url="https://www.superdry.com/mens/t-shirts",
            html=html,
            status_code=200,
            mode="requests",
            score=90,
            attempts=[FetchAttempt(mode="requests", url="https://www.superdry.com/mens/t-shirts", html=html, status_code=200)],
        )

        with patch("autonomous_crawler.runners.product_workflow.fetch_best_html", return_value=fetch):
            analysis = analyze_site_for_product_workflow("https://www.superdry.com/mens/t-shirts")

        discovery = analysis["extraction_contract_discovery"]
        context = analysis["extraction_context"]
        constraints = analysis["profile"]["constraints"]
        fields = {item["name"]: item for item in analysis["field_candidates"]}

        self.assertEqual(discovery["best_contract"]["parser_strategy"]["name"], "gtm_data_attribute_extractor")
        self.assertGreater(discovery["best_sample_count"], 0)
        self.assertTrue(context["can_execute_extract_from_contract"])
        self.assertEqual(
            constraints["extraction_contract"]["parser_strategy"]["name"],
            "gtm_data_attribute_extractor",
        )
        self.assertIn("extraction_evidence", constraints)
        self.assertTrue(fields["title"]["api_path"])
        self.assertTrue(fields["highest_price"]["api_path"])
        self.assertGreaterEqual(fields["title"]["confidence"], 0.5)

    def test_export_product_records_writes_json_with_field_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = ProductStore(root / "products.sqlite3")
            store.upsert_many([
                ProductRecord(
                    run_id="run-export",
                    source_site="shop.test",
                    source_url="https://shop.test/p1",
                    canonical_url="https://shop.test/p1",
                    title="Alpha",
                    highest_price=12.5,
                    colors=["Black"],
                    sizes=["M"],
                    description="Nice",
                    image_urls=["https://shop.test/a.jpg"],
                    category="Women>Products>Leggings",
                )
            ])
            output = root / "out.json"

            result = export_product_records(
                run_id="run-export",
                runtime_dir=str(root),
                export_spec=ExportSpec(
                    format="json",
                    output_path=str(output),
                    field_mapping={"title": "Title", "highest_price": "Price"},
                ),
            )

            rows = json.loads(output.read_text(encoding="utf-8"))
            self.assertEqual(result["record_count"], 1)
            self.assertEqual(rows[0], {"Title": "Alpha", "Price": 12.5})

    def test_export_product_records_rewrites_extension_to_format(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = ProductStore(root / "products.sqlite3")
            store.upsert_many([
                ProductRecord(
                    run_id="run-export-format",
                    title="Alpha",
                    canonical_url="https://shop.test/a",
                )
            ])
            output = root / "chosen.xlsx"

            result = export_product_records(
                run_id="run-export-format",
                runtime_dir=str(root),
                export_spec=ExportSpec(format="csv", output_path=str(output)),
            )

            self.assertTrue(result["output_path"].endswith(".csv"))
            self.assertTrue(Path(result["output_path"]).exists())


class ExportTemplateTests(unittest.TestCase):
    def test_from_dict_defaults(self) -> None:
        tmpl = ExportTemplate.from_dict(None)
        self.assertEqual(tmpl.sheet_name, "Sheet1")
        self.assertEqual(tmpl.start_row, 1)
        self.assertEqual(tmpl.start_column, 1)
        self.assertEqual(tmpl.field_to_column, {})
        self.assertEqual(tmpl.columns, [])

    def test_from_dict_full(self) -> None:
        data = {
            "sheet_name": "Products",
            "start_row": 3,
            "start_column": 2,
            "field_to_column": {"title": "Title", "highest_price": "Price"},
            "columns": ["title", "highest_price", "colors"],
        }
        tmpl = ExportTemplate.from_dict(data)
        self.assertEqual(tmpl.sheet_name, "Products")
        self.assertEqual(tmpl.start_row, 3)
        self.assertEqual(tmpl.start_column, 2)
        self.assertEqual(tmpl.field_to_column["title"], "Title")
        self.assertEqual(tmpl.columns, ["title", "highest_price", "colors"])

    def test_from_dict_clamps_start_values(self) -> None:
        tmpl = ExportTemplate.from_dict({"start_row": -5, "start_column": 0})
        self.assertEqual(tmpl.start_row, 1)
        self.assertEqual(tmpl.start_column, 1)

    def test_to_dict(self) -> None:
        tmpl = ExportTemplate(sheet_name="Data", start_row=2, field_to_column={"title": "T"})
        d = tmpl.to_dict()
        self.assertEqual(d["sheet_name"], "Data")
        self.assertEqual(d["start_row"], 2)
        self.assertEqual(d["field_to_column"], {"title": "T"})

    def test_xlsx_template_writes_headers_at_offset(self) -> None:
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = ProductStore(root / "products.sqlite3")
            store.upsert_many([
                ProductRecord(
                    run_id="run-tmpl",
                    title="Widget",
                    highest_price=9.99,
                    colors=["Red"],
                    canonical_url="https://shop.test/w",
                )
            ])
            output = root / "out.xlsx"
            tmpl = ExportTemplate(
                sheet_name="MyData",
                start_row=5,
                start_column=3,
                field_to_column={"title": "Product Name", "highest_price": "Unit Price"},
            )
            export_product_records(
                run_id="run-tmpl",
                runtime_dir=str(root),
                export_spec=ExportSpec(format="xlsx", output_path=str(output), template=tmpl),
            )
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            self.assertEqual(ws.title, "MyData")
            # Header at row 5, col 3 and 4
            self.assertEqual(ws.cell(5, 3).value, "Product Name")
            self.assertEqual(ws.cell(5, 4).value, "Unit Price")
            # Data at row 6
            self.assertEqual(ws.cell(6, 3).value, "Widget")
            self.assertAlmostEqual(ws.cell(6, 4).value, 9.99)

    def test_xlsx_default_export_still_works(self) -> None:
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = ProductStore(root / "products.sqlite3")
            store.upsert_many([
                ProductRecord(
                    run_id="run-default",
                    title="Gadget",
                    highest_price=19.99,
                    canonical_url="https://shop.test/g",
                )
            ])
            output = root / "default.xlsx"
            export_product_records(
                run_id="run-default",
                runtime_dir=str(root),
                export_spec=ExportSpec(format="xlsx", output_path=str(output)),
            )
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            # Default export: dict key order preserved, "title" is first
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            self.assertIn("title", headers)
            title_col = headers.index("title") + 1
            self.assertEqual(ws.cell(2, title_col).value, "Gadget")

    def test_xlsx_template_columns_list_controls_order(self) -> None:
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = ProductStore(root / "products.sqlite3")
            store.upsert_many([
                ProductRecord(
                    run_id="run-cols",
                    title="Item",
                    highest_price=5.0,
                    canonical_url="https://shop.test/i",
                )
            ])
            output = root / "cols.xlsx"
            tmpl = ExportTemplate(columns=["title", "highest_price"])
            export_product_records(
                run_id="run-cols",
                runtime_dir=str(root),
                export_spec=ExportSpec(format="xlsx", output_path=str(output), template=tmpl),
            )
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            self.assertEqual(ws.cell(1, 1).value, "title")
            self.assertEqual(ws.cell(1, 2).value, "highest_price")
            self.assertEqual(ws.cell(2, 1).value, "Item")
            self.assertAlmostEqual(ws.cell(2, 2).value, 5.0)


class StatusHelperTests(unittest.TestCase):
    def test_derive_current_stage_completed(self) -> None:
        self.assertEqual(_derive_current_stage("completed", 0, 10, 0, 10), "finished")

    def test_derive_current_stage_failed(self) -> None:
        self.assertEqual(_derive_current_stage("failed", 0, 5, 3, 8), "stopped")

    def test_derive_current_stage_running_with_queue(self) -> None:
        self.assertEqual(_derive_current_stage("running", 10, 0, 0, 10), "crawling")

    def test_derive_current_stage_running_finishing(self) -> None:
        self.assertEqual(_derive_current_stage("running", 0, 10, 0, 10), "finishing")

    def test_derive_current_stage_running_starting(self) -> None:
        self.assertEqual(_derive_current_stage("running", 0, 0, 0, 0), "starting")

    def test_extract_last_error_from_job(self) -> None:
        job = {"error": "connection timeout"}
        self.assertEqual(_extract_last_error_snippet(job, {}), "connection timeout")

    def test_extract_last_error_from_failures(self) -> None:
        profile_run = {"failures": [{"error": "parse failed"}]}
        self.assertEqual(_extract_last_error_snippet({}, profile_run), "parse failed")

    def test_extract_last_error_truncates(self) -> None:
        job = {"error": "x" * 300}
        self.assertEqual(len(_extract_last_error_snippet(job, {})), 200)

    def test_extract_last_error_empty(self) -> None:
        self.assertEqual(_extract_last_error_snippet({}, {}), "")

    def test_build_progress_summary_completed(self) -> None:
        s = _build_progress_summary("completed", 50, 50, 0, 0, 1.0)
        self.assertEqual(s, "Done — 50 records saved")

    def test_build_progress_summary_failed(self) -> None:
        s = _build_progress_summary("failed", 10, 7, 3, 0, 0.7)
        self.assertEqual(s, "Failed after 10 records (3 errors)")

    def test_build_progress_summary_running(self) -> None:
        s = _build_progress_summary("running", 20, 15, 2, 5, 0.75)
        self.assertIn("75%", s)
        self.assertIn("20 saved", s)
        self.assertIn("2 failed", s)
        self.assertIn("5 queued", s)

    def test_derive_quality_indicator_pass(self) -> None:
        self.assertEqual(_derive_quality_indicator({"field_coverage": 0.95}, 100, 0), "pass")

    def test_derive_quality_indicator_warn(self) -> None:
        self.assertEqual(_derive_quality_indicator({"field_coverage": 0.7}, 70, 10), "warn")

    def test_derive_quality_indicator_fail(self) -> None:
        self.assertEqual(_derive_quality_indicator({"field_coverage": 0.3}, 30, 50), "fail")

    def test_derive_quality_indicator_from_success_rate(self) -> None:
        self.assertEqual(_derive_quality_indicator({}, 95, 5), "pass")
        self.assertEqual(_derive_quality_indicator({}, 70, 30), "warn")
        self.assertEqual(_derive_quality_indicator({}, 30, 70), "fail")

    def test_derive_quality_indicator_unknown(self) -> None:
        self.assertEqual(_derive_quality_indicator({}, 0, 0), "unknown")


class StatusEndpointTests(unittest.TestCase):
    def setUp(self) -> None:
        _clear_jobs()

    def test_status_includes_new_fields(self) -> None:
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "accepted": True,
                "status": "completed",
                "run_id": "run-status",
                "product_stats": {"total": 5},
                "runner_summary": {"claimed": 5, "records_saved": 5, "succeeded": 5},
                "frontier_stats": {"done": 5},
                "quality_summary": {"field_coverage": 0.92},
            }
            client = TestClient(create_app())
            resp = client.post(
                "/runs/test",
                json={
                    "target_url": "https://shop.test",
                    "profile": {"name": "shop-status", "crawl_preferences": {"seed_urls": ["https://shop.test/c"]}},
                    "selected_fields": ["title"],
                },
            )
            task_id = resp.json()["task_id"]
            status = client.get(f"/runs/{task_id}/status")
            for _ in range(30):
                if status.json().get("status") == "completed":
                    break
                time.sleep(0.05)
                status = client.get(f"/runs/{task_id}/status")

        self.assertEqual(status.status_code, 200)
        body = status.json()
        self.assertEqual(body["current_stage"], "finished")
        self.assertEqual(body["quality_indicator"], "pass")
        self.assertIn("progress_summary", body)
        self.assertIn("last_error", body)
        self.assertIn("managed_ai", body)
        self.assertIn("ai_decisions", body)
        self.assertIn("llm_traces", body)
        self.assertIn("ai_repair_suggestions", body)
        self.assertIn("ai_patch_applications", body)
        self.assertIn("supervision", body)

    def test_status_and_events_include_runtime_supervision(self) -> None:
        supervision = {
            "enabled": True,
            "event_count": 1,
            "highest_severity": "critical",
            "last_event": {
                "action": "pause",
                "reason": "2 consecutive batches produced no records",
                "severity": "critical",
            },
            "recommended_next_action": "ai_rerun",
        }
        with patch("autonomous_crawler.api.routers.runs.build_advisor_from_config") as mock_build, \
                patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            advisor = MagicMock()
            advisor.provider = "test-provider"
            advisor.model = "test-model"
            advisor.review_run_plan.return_value = {}
            advisor.diagnose_run_result.return_value = {}
            mock_build.return_value = advisor
            mock_run.return_value = {
                "accepted": False,
                "status": "paused",
                "run_id": "run-supervised-status",
                "product_stats": {"total": 0},
                "runner_summary": {"claimed": 2, "records_saved": 0, "succeeded": 2},
                "frontier_stats": {"done": 2, "queued": 1},
                "diagnostics": {"supervision": supervision},
            }
            client = TestClient(create_app())
            resp = client.post(
                "/runs/test",
                json={
                    "target_url": "https://shop.test",
                    "profile": {"name": "shop-status", "crawl_preferences": {"seed_urls": ["https://shop.test/c"]}},
                    "selected_fields": ["title"],
                    "managed_ai": {"enabled": True, "mode": "full_managed"},
                    "llm": {"enabled": True, "base_url": "https://llm.example/v1", "model": "m"},
                },
            )
            task_id = resp.json()["task_id"]
            status = client.get(f"/runs/{task_id}/status")
            events = client.get(f"/runs/{task_id}/events")
            for _ in range(30):
                if status.json().get("supervision"):
                    break
                time.sleep(0.05)
                status = client.get(f"/runs/{task_id}/status")
                events = client.get(f"/runs/{task_id}/events")

        self.assertEqual(status.status_code, 200)
        self.assertEqual(status.json()["supervision"]["recommended_next_action"], "ai_rerun")
        self.assertTrue(any(item["type"] == "supervision_pause" for item in events.json()["events"]))


class ManagedAIRunTests(unittest.TestCase):
    def setUp(self) -> None:
        _clear_jobs()

    def test_product_run_without_managed_ai_has_empty_ai_state(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-no-ai",
                "status": "completed",
                "accepted": True,
                "product_stats": {"total": 0},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {"name": "shop.test", "crawl_preferences": {"seed_urls": ["https://shop.test/"]}},
                "selected_fields": ["title"],
                "test_limit": 1,
            })
        self.assertEqual(response.status_code, 200)
        task_id = response.json()["task_id"]

        status = {}
        for _ in range(20):
            status = client.get(f"/runs/{task_id}/status").json()
            if status["status"] == "completed":
                break
            time.sleep(0.05)

        self.assertFalse(status["managed_ai"]["enabled"])
        self.assertEqual(status["ai_decisions"], [])
        self.assertEqual(status["ai_repair_suggestions"], [])

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_managed_ai_pre_and_post_decisions_are_queryable(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.provider = "test-provider"
        advisor.model = "test-model"
        advisor.review_run_plan.return_value = {
            "approved": True,
            "risk_level": "low",
            "reasoning_summary": "Plan has usable seed URLs.",
            "warnings": [],
            "recommended_actions": ["Run bounded test first"],
        }
        advisor.diagnose_run_result.return_value = {
            "status_assessment": "good",
            "reasoning_summary": "Run produced expected records.",
            "likely_causes": [],
            "repair_suggestions": [
                {"action": "Continue with full run", "priority": "low", "rationale": "Quality is acceptable"}
            ],
        }
        mock_build.return_value = advisor
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-ai",
                "status": "completed",
                "accepted": True,
                "product_stats": {"total": 5},
                "quality": {"quality_gate": {"severity": "pass"}},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/"], "seed_kind": "list"},
                },
                "selected_fields": ["title"],
                "test_limit": 1,
                "llm": {
                    "enabled": True,
                    "base_url": "https://llm.example/v1",
                    "model": "test-model",
                    "api_key": "sk-test",
                },
                "managed_ai": {
                    "enabled": True,
                    "mode": "supervised",
                    "pre_run_review": True,
                    "post_run_diagnosis": True,
                },
            })
        self.assertEqual(response.status_code, 200)
        task_id = response.json()["task_id"]

        status = {}
        for _ in range(20):
            status = client.get(f"/runs/{task_id}/status").json()
            if status["status"] == "completed":
                break
            time.sleep(0.05)

        self.assertTrue(status["managed_ai"]["enabled"])
        self.assertEqual(status["managed_ai"]["mode"], "supervised")
        stages = [item["stage"] for item in status["ai_decisions"]]
        self.assertIn("pre_run_review", stages)
        self.assertIn("post_run_diagnosis", stages)
        trace_stages = [item["stage"] for item in status["llm_traces"]]
        self.assertIn("pre_run_review", trace_stages)
        self.assertIn("post_run_diagnosis", trace_stages)
        self.assertEqual(status["llm_traces"][0]["status"], "ok")
        self.assertIn("input_summary", status["llm_traces"][0])
        self.assertEqual(status["ai_diagnostics"]["status_assessment"], "good")
        self.assertEqual(status["ai_repair_suggestions"][0]["action"], "Continue with full run")

        events = client.get(f"/runs/{task_id}/events").json()["events"]
        event_types = [event["type"] for event in events]
        self.assertIn("ai_pre_run_review", event_types)
        self.assertIn("ai_post_run_diagnosis", event_types)
        self.assertIn("llm_trace_pre_run_review", event_types)
        self.assertIn("llm_trace_post_run_diagnosis", event_types)

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_managed_ai_can_apply_allowlisted_pre_run_profile_patch(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.provider = "test-provider"
        advisor.model = "test-model"
        advisor.review_run_plan.return_value = {
            "approved": True,
            "risk_level": "medium",
            "reasoning_summary": "Switch to dynamic mode and use corrected seeds.",
            "profile_patch": {
                "crawl_preferences": {
                    "seed_urls": ["https://shop.test/corrected"],
                    "seed_kind": "list",
                    "max_items": 25,
                },
                "access_config": {
                    "mode": "dynamic",
                    "wait_until": "networkidle",
                    "browser_config": {"capture_api": True, "render_time_ms": 3000},
                },
                "selectors": {"title": "h1.product-title"},
                "pagination_hints": {"type": "dom_links", "next_selector": "a.next"},
                "quality_expectations": {"required_fields": ["title", "highest_price"]},
            },
        }
        advisor.diagnose_run_result.return_value = {}
        mock_build.return_value = advisor
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-ai-patch",
                "status": "completed",
                "accepted": True,
                "product_stats": {"total": 1},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/old"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
                "test_limit": 1,
                "llm": {
                    "enabled": True,
                    "base_url": "https://llm.example/v1",
                    "model": "test-model",
                },
                "managed_ai": {
                    "enabled": True,
                    "mode": "supervised",
                    "pre_run_review": True,
                    "apply_pre_run_patch": True,
                },
            })
        self.assertEqual(response.status_code, 200)
        task_id = response.json()["task_id"]

        status = {}
        for _ in range(20):
            status = client.get(f"/runs/{task_id}/status").json()
            if status["status"] == "completed":
                break
            time.sleep(0.05)

        request = mock_run.call_args.args[0]
        profile = request.profile
        self.assertEqual(profile["crawl_preferences"]["seed_urls"], ["https://shop.test/corrected"])
        self.assertEqual(profile["crawl_preferences"]["max_items"], 25)
        self.assertEqual(profile["access_config"]["mode"], "dynamic")
        self.assertEqual(profile["selectors"]["title"], "h1.product-title")
        self.assertTrue(status["managed_ai"]["apply_pre_run_patch"])
        self.assertTrue(status["ai_patch_applications"][0]["applied"])
        self.assertIn("crawl_preferences.seed_urls", status["ai_patch_applications"][0]["accepted"])

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_managed_ai_can_apply_api_replay_profile_patch(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.provider = "test-provider"
        advisor.model = "test-model"
        advisor.review_run_plan.return_value = {
            "approved": True,
            "risk_level": "low",
            "reasoning_summary": "Promote observed product API.",
            "profile_patch": {
                "api_hints": {
                    "endpoint": "https://shop.test/api/products",
                    "method": "GET",
                    "format": "json",
                    "items_path": "data.items",
                    "field_mapping": {"title": "name", "highest_price": "price"},
                },
                "pagination_hints": {
                    "type": "page",
                    "page_param": "page",
                    "page_size_param": "limit",
                    "page_size": 20,
                },
                "crawl_preferences": {
                    "seed_kind": "api",
                    "seed_urls": ["https://shop.test/api/products"],
                },
            },
        }
        advisor.diagnose_run_result.return_value = {}
        mock_build.return_value = advisor
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-api-patch",
                "status": "completed",
                "accepted": True,
                "product_stats": {"total": 1},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/list",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                },
                "selected_fields": ["title", "highest_price"],
                "test_limit": 1,
                "llm": {
                    "enabled": True,
                    "base_url": "https://llm.example/v1",
                    "model": "test-model",
                },
                "managed_ai": {
                    "enabled": True,
                    "mode": "supervised",
                    "pre_run_review": True,
                    "apply_pre_run_patch": True,
                },
            })
        self.assertEqual(response.status_code, 200)
        task_id = response.json()["task_id"]

        status = {}
        for _ in range(20):
            status = client.get(f"/runs/{task_id}/status").json()
            if status["status"] == "completed":
                break
            time.sleep(0.05)

        request = mock_run.call_args.args[0]
        profile = request.profile
        self.assertEqual(profile["api_hints"]["endpoint"], "https://shop.test/api/products")
        self.assertEqual(profile["api_hints"]["items_path"], "data.items")
        self.assertEqual(profile["api_hints"]["field_mapping"]["title"], "name")
        self.assertEqual(profile["pagination_hints"]["type"], "page")
        self.assertEqual(profile["crawl_preferences"]["seed_kind"], "api")
        self.assertIn("api_hints.endpoint", status["ai_patch_applications"][0]["accepted"])
        self.assertIn("api_hints.field_mapping", status["ai_patch_applications"][0]["accepted"])

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_managed_ai_can_apply_post_graphql_replay_profile_patch(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.provider = "test-provider"
        advisor.model = "test-model"
        advisor.review_run_plan.return_value = {
            "approved": True,
            "risk_level": "low",
            "reasoning_summary": "Replay captured GraphQL product listing.",
            "profile_patch": {
                "api_hints": {
                    "endpoint": "https://shop.test/graphql",
                    "method": "POST",
                    "format": "graphql",
                    "kind": "graphql",
                    "items_path": "data.products.items",
                    "field_mapping": {"title": "name", "highest_price": "price"},
                    "headers": {
                        "content-type": "application/json",
                        "x-store": "nl",
                    },
                    "post_json": {
                        "operationName": "CategoryProducts",
                        "query": "query CategoryProducts { products { items { name price } } }",
                        "variables": {"currentPage": 1, "pageSize": 24},
                    },
                    "replay_diagnostics": {
                        "schema_version": "replay-diagnostics/v1",
                        "replay_required": True,
                        "risk_level": "low",
                        "dynamic_inputs": [
                            {
                                "name": "requestId",
                                "location": "json",
                                "path": "variables.requestId",
                                "generation_method": "random_hex_16",
                                "refresh_each_request": True,
                                "required": True,
                            }
                        ],
                        "signed_components": [
                            {"location": "header", "name": "x-signature", "kind": "signature_or_token"}
                        ],
                    },
                    "replay_runtime": {
                        "hook_name": "api_request_signature",
                        "secret_key": "fixture-secret",
                        "output_bindings": [
                            {
                                "source": "api_request_signature",
                                "location": "header",
                                "path": "x-signature",
                                "value_type": "hook",
                            }
                        ],
                    },
                },
                "pagination_hints": {
                    "type": "page",
                    "json_page_path": "variables.currentPage",
                    "json_page_size_path": "variables.pageSize",
                    "page_size": 24,
                },
                "crawl_preferences": {
                    "seed_kind": "api",
                    "seed_urls": ["https://shop.test/graphql"],
                },
            },
        }
        advisor.diagnose_run_result.return_value = {}
        mock_build.return_value = advisor
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-graphql-patch",
                "status": "completed",
                "accepted": True,
                "product_stats": {"total": 1},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/list",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                },
                "selected_fields": ["title", "highest_price"],
                "test_limit": 1,
                "llm": {
                    "enabled": True,
                    "base_url": "https://llm.example/v1",
                    "model": "test-model",
                },
                "managed_ai": {
                    "enabled": True,
                    "mode": "supervised",
                    "pre_run_review": True,
                    "apply_pre_run_patch": True,
                },
            })
        self.assertEqual(response.status_code, 200)
        task_id = response.json()["task_id"]

        status = {}
        for _ in range(20):
            status = client.get(f"/runs/{task_id}/status").json()
            if status["status"] == "completed":
                break
            time.sleep(0.05)

        request = mock_run.call_args.args[0]
        profile = request.profile
        self.assertEqual(profile["api_hints"]["endpoint"], "https://shop.test/graphql")
        self.assertEqual(profile["api_hints"]["method"], "POST")
        self.assertEqual(profile["api_hints"]["format"], "graphql")
        self.assertEqual(profile["api_hints"]["kind"], "graphql")
        self.assertEqual(profile["api_hints"]["headers"]["x-store"], "nl")
        self.assertEqual(profile["api_hints"]["post_json"]["variables"]["currentPage"], 1)
        self.assertTrue(profile["api_hints"]["replay_diagnostics"]["replay_required"])
        self.assertEqual(profile["api_hints"]["replay_runtime"]["hook_name"], "api_request_signature")
        self.assertEqual(profile["pagination_hints"]["json_page_path"], "variables.currentPage")
        self.assertEqual(profile["pagination_hints"]["json_page_size_path"], "variables.pageSize")
        accepted = status["ai_patch_applications"][0]["accepted"]
        self.assertIn("api_hints.headers", accepted)
        self.assertIn("api_hints.post_json", accepted)
        self.assertIn("api_hints.replay_diagnostics", accepted)
        self.assertIn("api_hints.replay_runtime", accepted)
        self.assertIn("pagination_hints.json_page_path", accepted)

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_managed_ai_rejects_unsafe_pre_run_profile_patch_values(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.provider = "test-provider"
        advisor.model = "test-model"
        advisor.review_run_plan.return_value = {
            "approved": True,
            "reasoning_summary": "Patch contains unsafe values.",
            "profile_patch": {
                "crawl_preferences": {"seed_urls": ["javascript:alert(1)"]},
                "access_config": {"mode": "shell"},
                "selectors": {"title": "h1\n<script>"},
            },
        }
        mock_build.return_value = advisor
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-ai-reject",
                "status": "completed",
                "accepted": True,
                "product_stats": {"total": 1},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/original"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
                "llm": {
                    "enabled": True,
                    "base_url": "https://llm.example/v1",
                    "model": "test-model",
                },
                "managed_ai": {
                    "enabled": True,
                    "mode": "supervised",
                    "pre_run_review": True,
                    "apply_pre_run_patch": True,
                },
            })
        self.assertEqual(response.status_code, 200)
        task_id = response.json()["task_id"]

        status = {}
        for _ in range(20):
            status = client.get(f"/runs/{task_id}/status").json()
            if status["status"] == "completed":
                break
            time.sleep(0.05)

        request = mock_run.call_args.args[0]
        profile = request.profile
        self.assertEqual(profile["crawl_preferences"]["seed_urls"], ["https://shop.test/original"])
        self.assertEqual(profile["access_config"]["mode"], "static")
        self.assertFalse(status["ai_patch_applications"][0]["applied"])
        self.assertIn("crawl_preferences.seed_urls", status["ai_patch_applications"][0]["rejected"])

    def test_managed_ai_requires_enabled_llm(self) -> None:
        client = TestClient(create_app())
        response = client.post("/runs/test", json={
            "target_url": "https://shop.test/",
            "profile": {"name": "shop.test", "crawl_preferences": {"seed_urls": ["https://shop.test/"]}},
            "managed_ai": {"enabled": True, "mode": "supervised"},
        })
        self.assertEqual(response.status_code, 400)
        self.assertIn("managed_ai requires llm.enabled=true", response.text)

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_ai_rerun_turns_diagnostics_into_executable_child_run(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.provider = "test-provider"
        advisor.model = "test-model"
        advisor.diagnose_run_result.return_value = {
            "status_assessment": "needs_attention",
            "reasoning_summary": "Static run missed product cards; retry dynamically.",
            "repair_suggestions": [
                {"action": "Retry with browser mode and corrected selector", "priority": "high", "rationale": "No records"}
            ],
            "next_run_overrides": {
                "access_config": {"mode": "dynamic", "wait_until": "networkidle"},
                "selectors": {"title": "h1.product-title"},
                "selected_fields": ["title", "colors"],
                "item_workers": 7,
                "export": {"format": "csv"},
            },
        }
        mock_build.return_value = advisor
        client = TestClient(create_app())

        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-ai-rerun",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {"claimed": 1, "records_saved": 0, "failed": 1},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/old"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                    "selectors": {"title": ".old-title"},
                },
                "selected_fields": ["title"],
                "test_limit": 10,
                "llm": {
                    "enabled": True,
                    "base_url": "https://llm.example/v1",
                    "model": "test-model",
                },
                "managed_ai": {
                    "enabled": True,
                    "mode": "analysis_only",
                    "post_run_diagnosis": True,
                },
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)

            rerun = client.post(f"/runs/{task_id}/ai-rerun", json={"run_kind": "test"})
            self.assertEqual(rerun.status_code, 200)
            child_id = rerun.json()["task_id"]
            for _ in range(20):
                child_status = client.get(f"/runs/{child_id}/status").json()
                if child_status["status"] == "completed":
                    break
                time.sleep(0.05)

        child_request = mock_run.call_args_list[-1].args[0]
        child_profile = child_request.profile
        self.assertEqual(child_request.item_workers, 7)
        self.assertEqual(child_profile["access_config"]["mode"], "dynamic")
        self.assertEqual(child_profile["access_config"]["wait_until"], "networkidle")
        self.assertEqual(child_profile["selectors"]["title"], "h1.product-title")
        self.assertEqual(child_profile["target_fields"], ["title", "colors"])
        self.assertEqual(child_status["parent_task_id"], task_id)
        self.assertEqual(child_status["repair_source"], "ai_rerun")
        self.assertTrue(child_status["ai_patch_applications"][0]["applied"])
        self.assertIn("profile.access_config.mode", child_status["ai_patch_applications"][0]["accepted"])

    def test_ai_rerun_rejects_invalid_override_values(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-ai-rerun-invalid",
                "status": "completed",
                "accepted": True,
                "product_stats": {"total": 1},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)

            rerun = client.post(f"/runs/{task_id}/ai-rerun", json={
                "extra_overrides": {
                    "access_config": {"mode": "shell"},
                    "selectors": {"title": "h1\nbad"},
                    "item_workers": 500,
                    "export": {"format": "exe"},
                }
            })
            self.assertEqual(rerun.status_code, 200)
            child_id = rerun.json()["task_id"]
            for _ in range(20):
                child_status = client.get(f"/runs/{child_id}/status").json()
                if child_status["status"] == "completed":
                    break
                time.sleep(0.05)

        child_request = mock_run.call_args_list[-1].args[0]
        child_profile = child_request.profile
        self.assertEqual(child_request.item_workers, 128)
        self.assertEqual(child_profile["access_config"]["mode"], "static")
        self.assertIn("profile.access_config.mode", child_status["ai_patch_applications"][0]["rejected"])
        self.assertIn("profile.selectors.title", child_status["ai_patch_applications"][0]["rejected"])
        self.assertNotIn("export.format", child_status["ai_patch_applications"][0]["accepted"])

    def test_ai_rerun_uses_supervision_repair_overrides(self) -> None:
        client = TestClient(create_app())
        supervision = {
            "enabled": True,
            "last_event": {
                "action": "pause",
                "reason": "2 consecutive batches produced no records or new URLs",
                "suggestions": [{"action": "switch_runtime_or_repair_selectors", "priority": "high"}],
            },
            "recommended_next_action": "ai_rerun",
        }
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-supervision-rerun",
                "status": "paused",
                "accepted": False,
                "product_stats": {"total": 0},
                "diagnostics": {"supervision": supervision},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "paused":
                    break
                time.sleep(0.05)

            rerun = client.post(f"/runs/{task_id}/ai-rerun", json={"run_kind": "test"})
            self.assertEqual(rerun.status_code, 200)
            child_id = rerun.json()["task_id"]
            for _ in range(20):
                child_status = client.get(f"/runs/{child_id}/status").json()
                if child_status["status"] == "paused":
                    break
                time.sleep(0.05)

        child_request = mock_run.call_args_list[-1].args[0]
        child_profile = child_request.profile
        self.assertEqual(child_profile["access_config"]["mode"], "dynamic")
        self.assertTrue(child_profile["access_config"]["browser_config"]["capture_api"])
        self.assertEqual(child_profile["access_config"]["wait_until"], "networkidle")
        self.assertIn("title", child_profile["selectors"])
        self.assertIn("profile.access_config.mode", child_status["ai_patch_applications"][0]["accepted"])

    def test_managed_actions_endpoint_executes_and_feeds_ai_rerun(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-managed-actions",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {"claimed": 1, "records_saved": 0},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)

            action_resp = client.post(f"/runs/{task_id}/managed-actions", json={"execute": True, "use_llm": False})
            self.assertEqual(action_resp.status_code, 200)
            action_body = action_resp.json()
            self.assertTrue(action_body["result"]["rerun_ready"])

            rerun = client.post(f"/runs/{task_id}/ai-rerun", json={"run_kind": "test"})
            self.assertEqual(rerun.status_code, 200)
            child_id = rerun.json()["task_id"]
            for _ in range(20):
                child_status = client.get(f"/runs/{child_id}/status").json()
                if child_status["status"] == "completed":
                    break
                time.sleep(0.05)

            events = client.get(f"/runs/{task_id}/events").json()["events"]

        child_request = mock_run.call_args_list[-1].args[0]
        child_profile = child_request.profile
        self.assertEqual(child_profile["access_config"]["mode"], "dynamic")
        self.assertTrue(child_profile["access_config"]["browser_config"]["capture_api"])
        self.assertTrue(any(item["type"] == "managed_actions_executed" for item in events))
        self.assertIn("profile.access_config.mode", child_status["ai_patch_applications"][0]["accepted"])

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_managed_actions_endpoint_accepts_llm_action_plan(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.choose_managed_actions.return_value = {
            "reasoning_summary": "Switch runtime and repair title.",
            "actions": [
                {"action": "adjust_runtime", "priority": "high", "params": {"mode": "dynamic"}},
                {"action": "repair_selectors", "priority": "high", "params": {"fields": ["title"]}},
                {"action": "prepare_rerun", "priority": "medium"},
            ],
        }
        mock_build.return_value = advisor
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-managed-actions-llm",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {"name": "shop.test", "crawl_preferences": {"seed_urls": ["https://shop.test/list"]}},
                "selected_fields": ["title"],
            })
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)
            action_resp = client.post(f"/runs/{task_id}/managed-actions", json={
                "execute": True,
                "use_llm": True,
                "llm": {"enabled": True, "base_url": "https://llm.example/v1", "model": "m"},
            })

        self.assertEqual(action_resp.status_code, 200)
        plan = action_resp.json()["result"]["plan"]
        self.assertEqual(plan["source"], "llm")
        self.assertEqual(plan["actions"][0]["action"], "adjust_runtime")
        self.assertEqual(plan["protocol_validation"]["schema_version"], "managed-action-plan/v2")
        advisor.choose_managed_actions.assert_called_once()

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_managed_actions_llm_receives_compact_managed_state(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.choose_managed_actions.return_value = {
            "reasoning_summary": "Need catalog repair.",
            "actions": [
                {"action": "select_catalog", "priority": "high", "params": {"target_url": "https://shop.test/list"}},
                {"action": "resolve_fields", "priority": "high", "params": {"fields": ["title", "colors"]}},
                {"action": "export_results", "priority": "low", "params": {"format": "csv"}},
            ],
        }
        mock_build.return_value = advisor
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-managed-actions-state",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {"name": "shop.test", "crawl_preferences": {"seed_urls": ["https://shop.test/list"]}},
                "selected_fields": ["title", "colors"],
            })
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)
            action_resp = client.post(f"/runs/{task_id}/managed-actions", json={
                "execute": False,
                "use_llm": True,
                "llm": {"enabled": True, "base_url": "https://llm.example/v1", "model": "m"},
            })

        self.assertEqual(action_resp.status_code, 200)
        call_kwargs = advisor.choose_managed_actions.call_args.kwargs
        self.assertIn("managed_llm_context", call_kwargs["diagnostics"])
        self.assertEqual(call_kwargs["run_spec"]["schema_version"], "managed-crawl-llm-context/v1")
        self.assertIn("workflow", call_kwargs["run_spec"])
        self.assertEqual(call_kwargs["diagnostics"]["managed_state"]["workflow"]["loop_name"], "AI Managed Crawl Loop v2")

    @patch("autonomous_crawler.api.routers.runs.build_advisor_from_config")
    def test_managed_actions_llm_can_drive_contract_extraction_from_extra_context(self, mock_build) -> None:
        advisor = MagicMock()
        advisor.choose_managed_actions.return_value = {
            "schema_version": "managed-action-plan/v2",
            "reasoning_summary": "A fixture extraction contract is available.",
            "actions": [
                {"action": "extract_from_contract", "priority": "high", "reason": "structured evidence is ready"},
            ],
        }
        mock_build.return_value = advisor
        contract = _read_extractor_fixture_json("superdry_com", "extraction_contract.json")
        html = _read_extractor_fixture_text("superdry_com", "raw_evidence_list_page.html")
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-managed-actions-contract",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://www.superdry.com/womens/tops",
                "profile": {
                    "name": "superdry",
                    "crawl_preferences": {"seed_urls": ["https://www.superdry.com/womens/tops"]},
                },
                "selected_fields": ["title", "highest_price", "color"],
            })
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)
            action_resp = client.post(f"/runs/{task_id}/managed-actions", json={
                "execute": True,
                "use_llm": True,
                "llm": {"enabled": True, "base_url": "https://llm.example/v1", "model": "m"},
                "extra_context": {
                    "extraction_contract": contract,
                    "extraction_evidence": html,
                    "source_url": "https://www.superdry.com/womens/tops",
                    "max_items": 3,
                },
            })
            status_after = client.get(f"/runs/{task_id}/status").json()
            events = client.get(f"/runs/{task_id}/events").json()["events"]

        self.assertEqual(action_resp.status_code, 200)
        body = action_resp.json()
        extraction = body["result"]["run_overrides"]["extraction_result"]
        self.assertEqual(extraction["item_count"], 3)
        self.assertEqual(extraction["site"], "superdry.com")
        self.assertEqual(extraction["items"][0]["title"], "Athletic Essentials Stripe Jersey Polo Shirt")
        self.assertTrue(body["result"]["rerun_ready"])
        self.assertEqual(status_after["managed_actions"][-1]["result"]["run_overrides"]["extraction_result"]["item_count"], 3)
        self.assertEqual(status_after["managed_state"]["extraction_context"]["latest_extraction"]["item_count"], 3)
        self.assertTrue(any(item["type"] == "managed_actions_executed" for item in events))
        call_kwargs = advisor.choose_managed_actions.call_args.kwargs
        self.assertTrue(call_kwargs["run_spec"]["extraction_context"]["can_execute_extract_from_contract"])
        self.assertEqual(call_kwargs["run_spec"]["extraction_context"]["parser_strategy"], "gtm_data_attribute_extractor")

    def test_analyzed_profile_feeds_contract_extraction_without_manual_extra_context(self) -> None:
        html = _read_extractor_fixture_text("superdry_com", "raw_evidence_list_page.html")
        fetch = BestFetchResult(
            url="https://www.superdry.com/womens/tops",
            html=html,
            status_code=200,
            mode="requests",
            score=90,
            attempts=[FetchAttempt(mode="requests", url="https://www.superdry.com/womens/tops", html=html, status_code=200)],
        )
        client = TestClient(create_app())

        with patch("autonomous_crawler.runners.product_workflow.fetch_best_html", return_value=fetch):
            analysis_resp = client.post("/site/analyze", json={
                "target_url": "https://www.superdry.com/womens/tops",
                "field_goal": "title price color image",
            })
        self.assertEqual(analysis_resp.status_code, 200)
        profile = analysis_resp.json()["profile"]

        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-analyzed-contract",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://www.superdry.com/womens/tops",
                "profile": profile,
                "selected_fields": ["title", "highest_price", "colors"],
                "test_limit": 5,
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)
            action_resp = client.post(f"/runs/{task_id}/managed-actions", json={
                "execute": True,
                "use_llm": False,
            })
            status_after = client.get(f"/runs/{task_id}/status").json()

        self.assertEqual(action_resp.status_code, 200)
        result = action_resp.json()["result"]
        self.assertEqual(result["plan"]["actions"][0]["action"], "extract_from_contract")
        extraction = result["run_overrides"]["extraction_result"]
        self.assertEqual(extraction["item_count"], 3)
        self.assertEqual(extraction["items"][0]["title"], "Athletic Essentials Stripe Jersey Polo Shirt")
        self.assertTrue(
            status_after["managed_state"]["extraction_context"]["can_execute_extract_from_contract"]
        )
        self.assertEqual(
            status_after["managed_state"]["extraction_context"]["parser_strategy"],
            "gtm_data_attribute_extractor",
        )

    def test_managed_repair_run_executes_actions_and_starts_child_run(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-managed-repair",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {"claimed": 1, "records_saved": 0},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
                "export": {"format": "xlsx", "output_path": "out.xlsx"},
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)

            repair = client.post(f"/runs/{task_id}/managed-repair-run", json={
                "execute": True,
                "use_llm": False,
                "run_kind": "test",
                "extra_context": {
                    "selected_fields": ["title", "colors"],
                    "export": {"format": "csv", "output_path": "fixed.csv"},
                },
            })

        self.assertEqual(repair.status_code, 200)
        body = repair.json()
        self.assertEqual(body["repair_source"], "managed_actions")
        self.assertIn("managed_action", body)
        self.assertTrue(body["managed_action"]["result"]["rerun_ready"])
        child_request = mock_run.call_args_list[-1].args[0]
        child_profile = child_request.profile
        self.assertEqual(child_profile["access_config"]["mode"], "dynamic")
        self.assertIn("colors", child_profile["selectors"]["detail"])
        self.assertEqual(child_profile["quality_expectations"]["required_fields"], ["title", "colors"])
        child_status = client.get(f"/runs/{body['task_id']}/status").json()
        self.assertEqual(child_status["ai_patch_applications"][0]["source"], "ai_diagnostics.next_run_overrides")

    def test_full_managed_auto_repair_starts_child_after_failed_quality(self) -> None:
        client = TestClient(create_app())
        calls = []

        def fake_run(request, *, task_id):
            calls.append((task_id, request))
            if len(calls) == 1:
                return {
                    "run_id": "run-auto-parent",
                    "status": "paused",
                    "accepted": False,
                    "product_stats": {"total": 0},
                    "runner_summary": {
                        "claimed": 2,
                        "records_saved": 0,
                        "succeeded": 2,
                        "supervision_events": [{
                            "action": "pause",
                            "reason": "2 consecutive batches produced no records",
                            "severity": "critical",
                        }],
                    },
                    "diagnostics": {
                        "supervision": {
                            "enabled": True,
                            "event_count": 1,
                            "last_event": {
                                "action": "pause",
                                "reason": "2 consecutive batches produced no records",
                                "severity": "critical",
                            },
                            "recommended_next_action": "ai_rerun",
                        }
                    },
                }
            return {
                "run_id": "run-auto-child",
                "status": "completed",
                "accepted": True,
                "product_stats": {"total": 3},
                "runner_summary": {"claimed": 3, "records_saved": 3, "succeeded": 3},
            }

        with patch("autonomous_crawler.api.routers.runs.build_advisor_from_config") as mock_build, \
                patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow", side_effect=fake_run):
            advisor = MagicMock()
            advisor.provider = "test-provider"
            advisor.model = "test-model"
            advisor.review_run_plan.return_value = {}
            advisor.diagnose_run_result.return_value = {
                "status_assessment": "failed",
                "reasoning_summary": "No records were extracted.",
                "next_run_overrides": {"access_config": {"mode": "dynamic"}},
            }
            advisor.choose_managed_actions.return_value = {
                "reasoning_summary": "Switch runtime and repair selectors.",
                "actions": [
                    {"action": "inspect_access", "priority": "high", "reason": "empty output"},
                    {"action": "repair_selectors", "priority": "high", "reason": "selectors failed"},
                    {"action": "prepare_rerun", "priority": "medium", "reason": "rerun"},
                ],
            }
            mock_build.return_value = advisor
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
                "llm": {
                    "enabled": True,
                    "base_url": "https://llm.example/v1",
                    "model": "test-model",
                },
                "managed_ai": {
                    "enabled": True,
                    "mode": "full_managed",
                    "auto_repair": True,
                    "post_run_diagnosis": True,
                },
            })

            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]

            status = {}
            for _ in range(30):
                status = client.get(f"/runs/{task_id}/status").json()
                auto_repair = status.get("managed_auto_repair") or {}
                if auto_repair.get("child_task_id"):
                    break
                time.sleep(0.05)

        self.assertTrue(status["managed_auto_repair"]["attempted"])
        child_id = status["managed_auto_repair"]["child_task_id"]
        self.assertTrue(child_id)
        self.assertEqual(status["managed_actions"][0]["result"]["plan"]["source"], "llm")
        child_status = client.get(f"/runs/{child_id}/status").json()
        self.assertEqual(child_status["parent_task_id"], task_id)
        self.assertEqual(child_status["managed_ai"]["mode"], "supervised")
        self.assertFalse(child_status["managed_ai"]["auto_repair"])
        events = client.get(f"/runs/{task_id}/events").json()["events"]
        self.assertTrue(any(item["type"] == "managed_auto_repair_started" for item in events))

    def test_managed_step_executes_actions_and_can_start_child_run(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-managed-step",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {"claimed": 2, "records_saved": 0, "failed": 2},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)

            step = client.post(f"/runs/{task_id}/managed-step", json={
                "execute": True,
                "use_llm": False,
                "start_child_run": True,
                "run_kind": "test",
                "extra_context": {
                    "selected_fields": ["title", "highest_price"],
                    "export": {"format": "csv", "output_path": "managed-step.csv"},
                },
            })

        self.assertEqual(step.status_code, 200)
        body = step.json()
        self.assertEqual(body["schema_version"], "managed-step/v1")
        self.assertEqual(body["stage"], "quality_review")
        self.assertTrue(body["action_record"]["result"]["rerun_ready"])
        self.assertTrue(body["child_run"]["task_id"])
        status = client.get(f"/runs/{task_id}/status").json()
        self.assertEqual(len(status["managed_steps"]), 1)
        events = client.get(f"/runs/{task_id}/events").json()["events"]
        self.assertTrue(any(item["type"] == "managed_step_executed" for item in events))

    def test_managed_control_loop_runs_probe_actions_and_child_rerun(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-control-loop",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {"claimed": 2, "records_saved": 0, "failed": 2},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title"],
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)

            with patch("autonomous_crawler.api.routers.runs._execute_inspect_access") as mock_probe:
                mock_probe.return_value = {
                    "action": "inspect_access",
                    "ok": True,
                    "summary": "runtime access knobs and evidence snapshot prepared",
                    "evidence": {
                        "target_url": "https://shop.test/",
                        "access_mode": "dynamic",
                        "browser_config": {"capture_api": True},
                        "sample_limit": 3,
                        "signals": ["xhr"],
                        "profile_summary": {"name": "shop.test", "target_fields": ["title"]},
                        "request": {"target_url": "https://shop.test/"},
                        "snapshot": {
                            "schema_version": "access-evidence/v1",
                            "summary": {"recommended_runtime": "dynamic_browser_probe"},
                        },
                        "base_snapshot": {
                            "schema_version": "access-evidence/v1",
                            "summary": {"recommended_runtime": "dynamic_browser_probe"},
                        },
                        "live_probe": False,
                    },
                    "patch": {"access_config": {"mode": "dynamic"}},
                    "overrides": {"access_config": {"mode": "dynamic"}},
                }
                loop = client.post(f"/runs/{task_id}/managed-control-loop", json={
                    "execute": True,
                    "use_llm": False,
                    "include_access_probe": True,
                    "start_child_run": True,
                    "run_kind": "test",
                    "extra_context": {
                        "selected_fields": ["title", "highest_price"],
                        "export": {"format": "csv", "output_path": "control-loop.csv"},
                    },
                })

        self.assertEqual(loop.status_code, 200)
        body = loop.json()
        self.assertEqual(body["schema_version"], "managed-control-loop/v1")
        self.assertEqual([item["stage"] for item in body["timeline"]], ["observe", "access_probe", "plan_act", "repair_rerun"])
        self.assertTrue(body["action_record"]["result"]["rerun_ready"])
        self.assertTrue(body["child_run"]["task_id"])
        status = client.get(f"/runs/{task_id}/status").json()
        self.assertEqual(status["latest_managed_control_loop"]["schema_version"], "managed-control-loop/v1")
        self.assertEqual(len(status["managed_control_loops"]), 1)
        self.assertEqual(len(status["access_probe_history"]), 1)
        events = client.get(f"/runs/{task_id}/events").json()["events"]
        self.assertTrue(any(item["type"] == "managed_control_loop_completed" for item in events))
        self.assertTrue(any(item["type"] == "access_probe_completed" for item in events))

    def test_managed_control_loop_can_promote_xhr_api_and_start_api_child_run(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-control-loop-api",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {"claimed": 1, "records_saved": 0, "failed": 1},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/list",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title", "highest_price"],
            })
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)

            with patch("autonomous_crawler.runners.managed_actions._should_collect_live_access_probe", return_value=True):
                with patch("autonomous_crawler.runners.managed_actions.NativeBrowserRuntime") as mock_runtime_cls:
                    mock_runtime = MagicMock()
                    mock_response = MagicMock()
                    mock_response.ok = True
                    mock_response.final_url = "https://shop.test/list"
                    mock_response.status_code = 200
                    mock_response.error = ""
                    mock_response.html = "<html><body></body></html>"
                    mock_response.to_dict.return_value = {
                        "captured_xhr": [{
                            "url": "https://shop.test/api/catalog?page=1&limit=20",
                            "method": "GET",
                            "status_code": 200,
                            "content_type": "application/json",
                            "body_preview": "{\"data\":{\"items\":[{\"name\":\"Alpha\",\"price\":10.5}]}}",
                        }],
                        "runtime_events": [],
                        "artifacts": [],
                        "engine_result": {"failure_classification": {"category": "none"}},
                    }
                    mock_runtime.render.return_value = mock_response
                    mock_runtime_cls.return_value = mock_runtime
                    loop = client.post(f"/runs/{task_id}/managed-control-loop", json={
                        "execute": True,
                        "use_llm": False,
                        "include_access_probe": True,
                        "live_probe": True,
                        "start_child_run": True,
                        "run_kind": "test",
                    })

        self.assertEqual(loop.status_code, 200)
        body = loop.json()
        self.assertTrue(body["action_record"]["result"]["profile_patch"]["api_hints"]["endpoint"])
        child_id = body["child_run"]["task_id"]
        child_status = client.get(f"/runs/{child_id}/status").json()
        patch_result = child_status["ai_patch_applications"][0]
        self.assertIn("profile.api_hints.endpoint", patch_result["accepted"])
        self.assertEqual(child_status["product_run_spec"]["profile"]["crawl_preferences"]["seed_kind"], "api")
        self.assertEqual(
            child_status["product_run_spec"]["profile"]["api_hints"]["endpoint"],
            "https://shop.test/api/catalog",
        )

    def test_status_and_managed_step_include_evidence_pack(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-evidence",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {
                    "claimed": 3,
                    "records_saved": 0,
                    "failed": 3,
                    "failure_buckets": {"challenge_like": 2, "timeout": 1},
                },
                "frontier_stats": {"done": 1, "failed": 2},
                "quality_summary": {
                    "total_records": 0,
                    "field_completeness": {"title": 0.0, "highest_price": 0.0},
                },
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title", "highest_price"],
            })
            task_id = response.json()["task_id"]
            for _ in range(20):
                status = client.get(f"/runs/{task_id}/status").json()
                if status["status"] == "completed":
                    break
                time.sleep(0.05)
            step = client.post(f"/runs/{task_id}/managed-step", json={"execute": False, "use_llm": False})

        self.assertEqual(status["evidence_pack"]["schema_version"], "run-evidence-pack/v1")
        self.assertEqual(status["evidence_pack"]["access_evidence"]["schema_version"], "access-evidence/v1")
        self.assertTrue(status["evidence_pack"]["access_evidence"]["summary"]["challenge_like"])
        self.assertEqual(status["evidence_pack"]["access_evidence"]["summary"]["recommended_runtime"], "protected_browser")
        self.assertIn("access_challenge", status["evidence_pack"]["recommended_focus"])
        self.assertEqual(status["evidence_pack"]["failure_evidence"]["failure_buckets"]["challenge_like"], 2)
        self.assertEqual(step.status_code, 200)
        self.assertEqual(step.json()["evidence_pack"]["quality_gaps"][0]["kind"], "field_coverage")
        self.assertEqual(step.json()["evidence_pack"]["access_evidence"]["schema_version"], "access-evidence/v1")

    def test_managed_state_endpoint_returns_unified_state_packet(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-managed-state",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {
                    "claimed": 2,
                    "records_saved": 0,
                    "failed": 2,
                    "failure_buckets": {"zero_records": 2},
                },
                "quality_summary": {
                    "total_records": 0,
                    "field_completeness": {"title": 0.0, "highest_price": 0.0},
                },
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "crawl_preferences": {"seed_urls": ["https://shop.test/list"], "seed_kind": "list"},
                    "access_config": {"mode": "static"},
                },
                "selected_fields": ["title", "highest_price"],
            })
        self.assertEqual(response.status_code, 200)
        task_id = response.json()["task_id"]

        status = {}
        for _ in range(20):
            status = client.get(f"/runs/{task_id}/status").json()
            if status["status"] == "completed":
                break
            time.sleep(0.05)

        managed_state = client.get(f"/runs/{task_id}/managed-state").json()
        state = managed_state["state"]
        llm_context = managed_state["llm_context"]

        self.assertEqual(managed_state["schema_version"], "managed-crawl-state/v1")
        self.assertEqual(state["task"]["task_id"], task_id)
        self.assertEqual(state["workflow"]["loop_name"], "AI Managed Crawl Loop v2")
        self.assertIn("zero_records", state["workflow"]["recommended_focus"])
        self.assertTrue(state["workflow"]["is_closed_loop_ready"])
        self.assertGreaterEqual(state["workflow"]["state_coverage"]["ready_count"], 6)
        self.assertEqual(state["progress"]["current_stage"], status["current_stage"])
        self.assertEqual(state["evidence_pack"]["schema_version"], "run-evidence-pack/v1")
        self.assertEqual(state["evidence_pack"]["quality_gaps"][0]["kind"], "field_coverage")
        self.assertGreater(len(state["timeline"]), 0)
        self.assertIn("recent_timeline", llm_context)
        self.assertEqual(llm_context["schema_version"], "managed-crawl-llm-context/v1")
        self.assertIn("workflow", llm_context)
        self.assertIn("quality_context", llm_context)

class ProductWorkflowAPITests(unittest.TestCase):
    def setUp(self) -> None:
        _clear_jobs()

    def test_catalog_import_endpoint(self) -> None:
        client = TestClient(create_app())
        response = client.post(
            "/catalog/import",
            json={"catalog": {"Women": {"Shoes": "https://shop.test/shoes"}}},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["schema_version"], "catalog-tree/v1")
        self.assertEqual(payload["leaf_count"], 1)

    def test_site_analyze_returns_catalog_fields_and_profile(self) -> None:
        client = TestClient(create_app())
        response = client.post(
            "/site/analyze",
            json={
                "target_url": "mock://catalog",
                "field_goal": "标题 价格 图片",
                "imported_catalog": {"Women": {"Jackets": "https://shop.test/jackets"}},
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["schema_version"], "site-analysis/v1")
        self.assertGreaterEqual(len(payload["field_candidates"]), 3)
        self.assertEqual(payload["catalog_tree"][0]["label"], "Women")
        self.assertIn("profile", payload)

    def test_fields_resolve_endpoint(self) -> None:
        client = TestClient(create_app())
        response = client.post(
            "/fields/resolve",
            json={
                "available_fields": [{"name": "title"}, {"name": "sizes"}],
                "natural_language": "我要商品标题和尺码",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["selected_fields"], ["sizes", "title"])

    def test_access_probe_endpoint_returns_snapshot_layers(self) -> None:
        client = TestClient(create_app())
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "run_id": "run-probe",
                "status": "completed",
                "accepted": False,
                "product_stats": {"total": 0},
                "runner_summary": {"claimed": 1, "records_saved": 0, "failed": 1},
                "frontier_stats": {"done": 1},
            }
            response = client.post("/runs/test", json={
                "target_url": "https://shop.test/",
                "profile": {
                    "name": "shop.test",
                    "access_config": {"mode": "dynamic"},
                },
                "selected_fields": ["title"],
            })
            task_id = response.json()["task_id"]

        with patch("autonomous_crawler.api.routers.runs._execute_inspect_access") as mock_probe:
            mock_probe.return_value = {
                "action": "inspect_access",
                "ok": True,
                "summary": "runtime access probe completed",
                "evidence": {
                    "target_url": "https://shop.test/",
                    "access_mode": "dynamic",
                    "browser_config": {"capture_api": True},
                    "sample_limit": 2,
                    "signals": ["xhr"],
                    "profile_summary": {"name": "shop.test", "target_fields": ["title"]},
                    "request": {"target_url": "https://shop.test/"},
                    "snapshot": {
                        "schema_version": "access-probe/v1",
                        "summary": {"challenge_like": False},
                        "probe_snapshot": {"schema_version": "access-probe/v1"},
                    },
                    "base_snapshot": {
                        "schema_version": "access-evidence/v1",
                        "summary": {"challenge_like": False},
                    },
                    "live_probe": True,
                },
                "patch": {"access_config": {"mode": "dynamic"}},
                "overrides": {"access_config": {"mode": "dynamic"}},
            }
            probe = client.post(f"/runs/{task_id}/access-probe", json={
                "target_url": "https://shop.test/",
                "live_probe": True,
                "sample_limit": 2,
            })

        self.assertEqual(probe.status_code, 200)
        body = probe.json()
        self.assertEqual(body["schema_version"], "access-probe-response/v1")
        self.assertEqual(body["snapshot"]["schema_version"], "access-probe/v1")
        self.assertEqual(body["base_snapshot"]["schema_version"], "access-evidence/v1")
        status = client.get(f"/runs/{task_id}/status").json()
        self.assertEqual(status["latest_access_probe"]["schema_version"], "access-probe-response/v1")
        self.assertEqual(len(status["access_probe_history"]), 1)

    def test_runs_test_registers_profile_job(self) -> None:
        with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow") as mock_run:
            mock_run.return_value = {
                "accepted": True,
                "status": "completed",
                "run_id": "run-test",
                "product_stats": {"total": 3},
                "runner_summary": {"claimed": 3, "records_saved": 3},
                "frontier_stats": {"done": 3},
            }
            client = TestClient(create_app())
            response = client.post(
                "/runs/test",
                json={
                    "target_url": "https://shop.test",
                    "profile": {"name": "shop-test", "crawl_preferences": {"seed_urls": ["https://shop.test/c"]}},
                    "catalog_nodes": [{"label": "C", "url": "https://shop.test/c", "path": ["C"]}],
                    "selected_fields": ["title", "highest_price"],
                    "item_workers": 4,
                    "test_limit": 100,
                    "runtime_dir": "dev_logs/runtime/test-api",
                },
            )
            self.assertEqual(response.status_code, 200)
            task_id = response.json()["task_id"]
            time.sleep(0.4)
            status = client.get(f"/runs/{task_id}/status")
            events = client.get(f"/runs/{task_id}/events")

        self.assertEqual(status.status_code, 200)
        self.assertEqual(status.json()["record_count"], 3)
        self.assertEqual(events.status_code, 200)
        self.assertGreaterEqual(len(events.json()["events"]), 2)
        request = mock_run.call_args.args[0]
        self.assertEqual(request.item_workers, 4)
        self.assertGreaterEqual(request.max_batches, 1)

    def test_exports_endpoint_uses_product_store(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = ProductStore(root / "products.sqlite3")
            store.upsert_many([
                ProductRecord(run_id="run-api-export", title="Beta", canonical_url="https://shop.test/b")
            ])
            output = root / "export.csv"
            client = TestClient(create_app())
            response = client.post(
                "/exports",
                json={
                    "run_id": "run-api-export",
                    "runtime_dir": str(root),
                    "format": "csv",
                    "output_path": str(output),
                },
            )
            self.assertEqual(response.status_code, 200)
            self.assertTrue(output.exists())
            self.assertEqual(response.json()["record_count"], 1)

    def test_product_run_auto_exports_after_completion(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            export_path = root / "chosen.xlsx"

            def fake_run(request, *, task_id):
                store = ProductStore(root / "products.sqlite3")
                store.upsert_many([
                    ProductRecord(run_id=request.run_id, title="Auto Export", canonical_url="https://shop.test/a")
                ])
                return {
                    "accepted": True,
                    "status": "completed",
                    "run_id": request.run_id,
                    "product_stats": {"total": 1},
                    "runner_summary": {"claimed": 1, "records_saved": 1},
                    "frontier_stats": {"done": 1},
                }

            with patch("autonomous_crawler.api.routers.runs.run_profile_longrun_workflow", side_effect=fake_run):
                client = TestClient(create_app())
                response = client.post(
                    "/runs/test",
                    json={
                        "target_url": "https://shop.test",
                        "profile": {"name": "shop-test", "crawl_preferences": {"seed_urls": ["https://shop.test/c"]}},
                        "catalog_nodes": [{"id": "c", "label": "C", "url": "https://shop.test/c", "path": ["C"]}],
                        "selected_fields": ["title"],
                        "item_workers": 1,
                        "test_limit": 10,
                        "runtime_dir": str(root),
                        "export": {"format": "csv", "output_path": str(export_path)},
                    },
                )
                self.assertEqual(response.status_code, 200)
                task_id = response.json()["task_id"]
                status = client.get(f"/runs/{task_id}/status")
                events = client.get(f"/runs/{task_id}/events")
                for _ in range(30):
                    if status.json().get("export"):
                        break
                    time.sleep(0.05)
                    status = client.get(f"/runs/{task_id}/status")
                    events = client.get(f"/runs/{task_id}/events")

            self.assertEqual(status.status_code, 200)
            body = status.json()
            self.assertEqual(body["record_count"], 1)
            self.assertEqual(body["export"]["format"], "csv")
            self.assertTrue(body["export"]["output_path"].endswith("chosen.csv"))
            self.assertTrue((root / "chosen.csv").exists())
            self.assertTrue(any(item["type"] == "export_ready" for item in events.json()["events"]))


if __name__ == "__main__":
    unittest.main()
