"""Product-facing workflow helpers for the CLM frontend/API layer.

This module is intentionally product-level glue. It turns crawler primitives
such as HTML recon, SiteProfile, and ProductStore into stable objects that a
frontend can render: catalog trees, field candidates, run specs, progress
events, and export artifacts.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import tempfile
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

from bs4 import BeautifulSoup
import httpx

from autonomous_crawler.llm import OpenAICompatibleAdvisor
from autonomous_crawler.models.product import ProductRecord
from autonomous_crawler.storage.product_store import ProductStore
from autonomous_crawler.tools.extraction_contracts import discover_extraction_contracts
from autonomous_crawler.tools.html_recon import build_recon_report, fetch_best_html

from .site_profile import SiteProfile


DEFAULT_PRODUCT_FIELDS = [
    "title",
    "highest_price",
    "colors",
    "sizes",
    "description",
    "image_urls",
    "canonical_url",
    "category_level_1",
    "category_level_2",
    "category_level_3",
]

MAX_ANALYSIS_EXTRACTION_EVIDENCE_CHARS = 500_000

FIELD_ALIASES = {
    "title": {"title", "name", "product_name", "商品标题", "标题", "名称"},
    "highest_price": {"highest_price", "price", "original_price", "最高价格", "原价", "价格", "吊牌价"},
    "colors": {"colors", "color", "colour", "颜色", "色号"},
    "sizes": {"sizes", "size", "尺码", "规格"},
    "description": {"description", "desc", "body", "详情", "描述", "商品描述"},
    "image_urls": {"image_urls", "images", "image", "main_image", "图片", "主图", "商品图"},
    "canonical_url": {"canonical_url", "url", "link", "商品链接", "详情页"},
    "category_level_1": {"category_level_1", "一级目录", "一级分类"},
    "category_level_2": {"category_level_2", "二级目录", "二级分类"},
    "category_level_3": {"category_level_3", "三级目录", "三级分类"},
}


@dataclass(frozen=True)
class CatalogNode:
    id: str
    label: str
    url: str = ""
    path: list[str] = field(default_factory=list)
    level1: str = ""
    level2: str = ""
    level3: str = ""
    children: list["CatalogNode"] = field(default_factory=list)
    source: str = "agent"

    def to_dict(self) -> dict[str, Any]:
        payload = {
            "id": self.id,
            "label": self.label,
            "url": self.url,
            "path": list(self.path),
            "level1": self.level1,
            "level2": self.level2,
            "level3": self.level3,
            "source": self.source,
            "children": [child.to_dict() for child in self.children],
        }
        if self.source.startswith("graphql:"):
            payload["source"] = "graphql"
            payload["graphql_category_uid"] = self.source.split(":", 1)[1]
        return payload


@dataclass(frozen=True)
class FieldCandidate:
    name: str
    label: str
    selected: bool = True
    source: str = "agent"
    selector: str = ""
    api_path: str = ""
    confidence: float = 0.5
    reason: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "label": self.label,
            "selected": self.selected,
            "source": self.source,
            "selector": self.selector,
            "api_path": self.api_path,
            "confidence": self.confidence,
            "reason": self.reason,
        }


@dataclass(frozen=True)
class ExportTemplate:
    """Controls xlsx output layout: sheet name, starting cell, field-to-column mapping."""
    sheet_name: str = "Sheet1"
    start_row: int = 1
    start_column: int = 1
    field_to_column: dict[str, str] = field(default_factory=dict)
    columns: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "start_row": self.start_row,
            "start_column": self.start_column,
            "field_to_column": dict(self.field_to_column),
            "columns": list(self.columns),
        }

    @classmethod
    def from_dict(cls, data: Any) -> "ExportTemplate":
        if not isinstance(data, dict):
            return cls()
        return cls(
            sheet_name=str(data.get("sheet_name") or "Sheet1"),
            start_row=max(1, int(data.get("start_row") or 1)),
            start_column=max(1, int(data.get("start_column") or 1)),
            field_to_column={str(k): str(v) for k, v in dict(data.get("field_to_column") or {}).items()},
            columns=[str(c) for c in (data.get("columns") or []) if str(c).strip()],
        )


@dataclass(frozen=True)
class ExportSpec:
    format: str = "xlsx"
    output_path: str = ""
    template_path: str = ""
    field_mapping: dict[str, str] = field(default_factory=dict)
    template: ExportTemplate = field(default_factory=ExportTemplate)


@dataclass(frozen=True)
class CrawlRunSpec:
    target_url: str
    profile: dict[str, Any]
    catalog_nodes: list[dict[str, Any]] = field(default_factory=list)
    selected_fields: list[str] = field(default_factory=list)
    export: ExportSpec = field(default_factory=ExportSpec)
    run_mode: str = "direct"
    item_workers: int = 4
    max_sites: int = 1
    test_limit: int = 100
    runtime_dir: str = ""


def import_catalog_tree(payload: Any, *, source: str = "import") -> list[dict[str, Any]]:
    """Normalize nested menu JSON into frontend-friendly catalog nodes."""
    nodes = _catalog_nodes_from_any(payload, path=[], source=source)
    return [node.to_dict() for node in nodes]


def analyze_site_for_product_workflow(
    url: str,
    *,
    imported_catalog: Any = None,
    field_goal: str = "",
    headers: dict[str, str] | None = None,
    advisor: OpenAICompatibleAdvisor | None = None,
) -> dict[str, Any]:
    """Analyze a site entry URL and return catalog/field/profile draft data."""
    fetch = fetch_best_html(url, headers=headers)
    html = fetch.html or ""
    final_url = fetch.url or url
    recon = build_recon_report(final_url, html) if html else _empty_recon(final_url)
    extraction_contract_discovery = _discover_analysis_extraction_contracts(
        html,
        final_url=final_url,
    )
    discovered_catalog = discover_catalog_tree(html, base_url=final_url)
    if not discovered_catalog:
        discovered_catalog = discover_catalog_from_site_fallbacks(final_url, html)
    imported_nodes = import_catalog_tree(imported_catalog, source="import") if imported_catalog else []
    llm_result = _run_site_analysis_advisor(
        advisor=advisor,
        target_url=final_url,
        field_goal=field_goal,
        recon=recon,
        discovered_catalog=discovered_catalog,
    )
    if llm_result.get("catalog_tree") and not imported_nodes:
        discovered_catalog = _catalog_from_llm_result(llm_result["catalog_tree"], fallback_source="llm") or discovered_catalog
    catalog_nodes = _enrich_catalog_with_discovered_metadata(imported_nodes or discovered_catalog, discovered_catalog)
    field_candidates = discover_field_candidates(recon, field_goal=field_goal)
    field_candidates = _merge_contract_field_candidates(
        field_candidates,
        extraction_contract_discovery,
    )
    field_candidates = _merge_llm_field_candidates(field_candidates, llm_result)
    profile = site_profile_from_analysis(
        url=final_url,
        recon=recon,
        catalog_nodes=catalog_nodes,
        field_candidates=field_candidates,
    )
    profile_data = profile.to_dict()
    profile_data = _attach_extraction_contract_to_profile(
        profile_data,
        extraction_contract_discovery,
        evidence=html,
        final_url=final_url,
    )
    if isinstance(llm_result.get("crawl_preferences"), dict):
        prefs = dict(profile_data.get("crawl_preferences") or {})
        llm_prefs = dict(llm_result.get("crawl_preferences") or {})
        if isinstance(llm_prefs.get("seed_urls"), list):
            existing = [str(item) for item in prefs.get("seed_urls") or [] if str(item).strip()]
            additions = [str(item) for item in llm_prefs.get("seed_urls") or [] if str(item).strip()]
            prefs["seed_urls"] = _dedupe_preserve_order([*existing, *additions])[:500]
        profile_data["crawl_preferences"] = prefs
    selectors = dict(profile_data.get("selectors") or {})
    llm_selectors = llm_result.get("selectors") if isinstance(llm_result.get("selectors"), dict) else {}
    for field in field_candidates:
      if field.selector:
        selectors[field.name] = field.selector
    for key, value in llm_selectors.items():
        selector = _safe_selector(value)
        if selector and _normalize_field_name(str(key)) == "item_container":
            selectors["item_container"] = selector
    if selectors:
        profile_data["selectors"] = selectors
    return {
        "schema_version": "site-analysis/v1",
        "target_url": url,
        "final_url": final_url,
        "status_code": fetch.status_code,
        "fetch_error": fetch.error,
        "catalog_tree": catalog_nodes,
        "discovered_catalog_tree": discovered_catalog,
        "imported_catalog_tree": imported_nodes,
        "field_candidates": [field.to_dict() for field in field_candidates],
        "profile": profile_data,
        "llm_analysis": llm_result,
        "extraction_contract_discovery": _public_extraction_contract_discovery(
            extraction_contract_discovery,
        ),
        "extraction_context": _analysis_extraction_context(
            extraction_contract_discovery,
            has_evidence=bool(html),
        ),
        "recon_summary": {
            "framework": recon.get("frontend_framework"),
            "rendering": recon.get("rendering"),
            "anti_bot": recon.get("anti_bot"),
            "item_count": recon.get("dom_structure", {}).get("item_count", 0),
            "field_selectors": recon.get("dom_structure", {}).get("field_selectors", {}),
            "product_selector": recon.get("dom_structure", {}).get("product_selector", ""),
        },
    }


def _discover_analysis_extraction_contracts(html: str, *, final_url: str) -> dict[str, Any]:
    if not html:
        return {
            "schema_version": "extraction-contract-discovery/v1",
            "source_url": final_url,
            "site": urlparse(final_url).netloc.lower().removeprefix("www."),
            "candidate_count": 0,
            "best_contract": None,
            "best_confidence": 0.0,
            "best_sample_count": 0,
            "candidates": [],
            "warnings": ["No HTML evidence available for extraction contract discovery."],
        }
    return discover_extraction_contracts(
        html,
        source_url=final_url,
        site=urlparse(final_url).netloc.lower().removeprefix("www."),
        sample_items=5,
    )


def _attach_extraction_contract_to_profile(
    profile_data: dict[str, Any],
    discovery: dict[str, Any],
    *,
    evidence: str,
    final_url: str,
) -> dict[str, Any]:
    best_contract = discovery.get("best_contract") if isinstance(discovery, dict) else None
    if not isinstance(best_contract, dict) or not best_contract:
        return profile_data
    updated = dict(profile_data or {})
    constraints = dict(updated.get("constraints") or {})
    constraints["extraction_contract"] = best_contract
    constraints["extraction_contract_discovery"] = _public_extraction_contract_discovery(discovery)
    constraints["extraction_evidence"] = str(evidence or "")[:MAX_ANALYSIS_EXTRACTION_EVIDENCE_CHARS]
    constraints["extraction_evidence_source_url"] = final_url
    constraints["extraction_evidence_type"] = "html"
    updated["constraints"] = constraints
    strategy = best_contract.get("parser_strategy") if isinstance(best_contract.get("parser_strategy"), dict) else {}
    updated["training_notes"] = _append_training_notes(
        updated.get("training_notes"),
        [
            (
                "extraction_contract:"
                f"{strategy.get('name', 'unknown')};"
                f"samples={int(discovery.get('best_sample_count') or 0)}"
            )
        ],
    )
    return updated


def _public_extraction_contract_discovery(discovery: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(discovery, dict):
        return {}

    def clean_candidate(candidate: Any) -> dict[str, Any]:
        if not isinstance(candidate, dict):
            return {}
        output = dict(candidate)
        samples = output.get("sample_items") if isinstance(output.get("sample_items"), list) else []
        output["sample_items"] = samples[:3]
        return output

    return {
        "schema_version": discovery.get("schema_version", "extraction-contract-discovery/v1"),
        "source_url": discovery.get("source_url", ""),
        "site": discovery.get("site", ""),
        "candidate_count": int(discovery.get("candidate_count") or 0),
        "best_contract": discovery.get("best_contract") if isinstance(discovery.get("best_contract"), dict) else None,
        "best_confidence": float(discovery.get("best_confidence") or 0.0),
        "best_sample_count": int(discovery.get("best_sample_count") or 0),
        "candidates": [
            item for item in (clean_candidate(candidate) for candidate in list(discovery.get("candidates") or [])[:6])
            if item
        ],
        "warnings": [str(item) for item in list(discovery.get("warnings") or [])[:10]],
    }


def _analysis_extraction_context(discovery: dict[str, Any], *, has_evidence: bool) -> dict[str, Any]:
    best_contract = discovery.get("best_contract") if isinstance(discovery, dict) else None
    strategy = best_contract.get("parser_strategy") if isinstance(best_contract, dict) and isinstance(best_contract.get("parser_strategy"), dict) else {}
    return {
        "schema_version": "analysis-extraction-context/v1",
        "has_contract": bool(best_contract),
        "parser_strategy": strategy.get("name", "") if strategy else "",
        "recommended_runtime": best_contract.get("recommended_clm_runtime", "") if isinstance(best_contract, dict) else "",
        "sample_count": int(discovery.get("best_sample_count") or 0) if isinstance(discovery, dict) else 0,
        "confidence": float(discovery.get("best_confidence") or 0.0) if isinstance(discovery, dict) else 0.0,
        "has_evidence": bool(has_evidence),
        "can_execute_extract_from_contract": bool(best_contract and has_evidence),
    }


def _merge_contract_field_candidates(
    candidates: list[FieldCandidate],
    discovery: dict[str, Any],
) -> list[FieldCandidate]:
    best_contract = discovery.get("best_contract") if isinstance(discovery, dict) else None
    if not isinstance(best_contract, dict):
        return candidates
    strategy = best_contract.get("parser_strategy") if isinstance(best_contract.get("parser_strategy"), dict) else {}
    field_paths = best_contract.get("field_paths") if isinstance(best_contract.get("field_paths"), dict) else {}
    if not field_paths:
        return candidates
    by_name = {field.name: field for field in candidates}
    confidence = max(0.55, min(float(discovery.get("best_confidence") or best_contract.get("confidence") or 0.55), 0.95))
    reason = f"available from extraction contract {strategy.get('name', 'unknown')}"
    for raw_name, raw_spec in field_paths.items():
        name = _normalize_contract_field_name(str(raw_name))
        if not name or name not in DEFAULT_PRODUCT_FIELDS:
            continue
        spec = raw_spec if isinstance(raw_spec, dict) else {}
        path = str(spec.get("path") or "")
        current = by_name.get(name)
        if current is None:
            by_name[name] = FieldCandidate(
                name=name,
                label=_field_label(name),
                selected=bool(spec.get("required", True)) or name in {"title", "highest_price"},
                source="contract",
                api_path=path,
                confidence=confidence,
                reason=reason,
            )
        else:
            by_name[name] = FieldCandidate(
                name=current.name,
                label=current.label,
                selected=current.selected,
                source="contract" if not current.selector else current.source,
                selector=current.selector,
                api_path=current.api_path or path,
                confidence=max(current.confidence, confidence),
                reason=current.reason or reason,
            )
    ordered: list[FieldCandidate] = []
    seen: set[str] = set()
    for field in candidates:
        ordered.append(by_name[field.name])
        seen.add(field.name)
    for name in DEFAULT_PRODUCT_FIELDS:
        if name in by_name and name not in seen:
            ordered.append(by_name[name])
            seen.add(name)
    for name, field in by_name.items():
        if name not in seen:
            ordered.append(field)
    return ordered


def _normalize_contract_field_name(name: str) -> str:
    aliases = {
        "color": "colors",
        "colour": "colors",
        "size": "sizes",
        "image_url": "image_urls",
        "images": "image_urls",
        "product_url": "canonical_url",
        "url": "canonical_url",
    }
    normalized = _normalize_field_name(name)
    return aliases.get(normalized, normalized)


def _run_site_analysis_advisor(
    *,
    advisor: OpenAICompatibleAdvisor | None,
    target_url: str,
    field_goal: str,
    recon: dict[str, Any],
    discovered_catalog: list[dict[str, Any]],
) -> dict[str, Any]:
    if advisor is None:
        return {"enabled": False}
    try:
        raw = advisor.analyze_site(target_url, field_goal, recon, discovered_catalog)
        if not isinstance(raw, dict):
            return {"enabled": True, "fallback_used": True, "error": "LLM site analysis did not return an object"}
        return {"enabled": True, "fallback_used": False, **raw}
    except Exception as exc:
        return {"enabled": True, "fallback_used": True, "error": str(exc)}


def _catalog_from_llm_result(value: Any, *, fallback_source: str) -> list[dict[str, Any]]:
    try:
        return import_catalog_tree(value, source=fallback_source)
    except Exception:
        return []


def _enrich_catalog_with_discovered_metadata(
    nodes: list[dict[str, Any]],
    discovered_nodes: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Copy engine-discovered metadata such as GraphQL category UID onto user/LLM nodes.

    Users may import a hand-written category tree that has the right URLs but
    lacks ecommerce backend metadata.  For Magento/PWA sites that metadata is
    the difference between slow empty browser-rendered category pages and the
    working GraphQL product API path.
    """
    if not nodes or not discovered_nodes:
        return nodes
    by_url: dict[str, dict[str, Any]] = {}
    by_path: dict[str, dict[str, Any]] = {}
    for item in _flatten_catalog_nodes(discovered_nodes):
        url_key = _catalog_url_key(str(item.get("url") or ""))
        path_key = _catalog_path_key(item)
        if url_key:
            by_url.setdefault(url_key, item)
        if path_key:
            by_path.setdefault(path_key, item)

    def visit(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        output: list[dict[str, Any]] = []
        for item in items or []:
            if not isinstance(item, dict):
                continue
            next_item = dict(item)
            match = by_url.get(_catalog_url_key(str(next_item.get("url") or ""))) or by_path.get(_catalog_path_key(next_item))
            if match:
                for key in ("graphql_category_uid", "source"):
                    if match.get(key) and not next_item.get(key):
                        next_item[key] = match[key]
            next_item["children"] = visit(list(next_item.get("children") or []))
            output.append(next_item)
        return output

    return visit(nodes)


def _catalog_url_key(url: str) -> str:
    parsed = urlparse(str(url or ""))
    if not parsed.netloc:
        return ""
    path = parsed.path.strip("/").lower()
    if path.endswith(".html"):
        path = path[:-5]
    return f"{parsed.netloc.lower()}/{path}".rstrip("/")


def _catalog_path_key(node: dict[str, Any]) -> str:
    path = node.get("path") if isinstance(node.get("path"), list) else []
    parts = [str(part).strip().lower() for part in path if str(part).strip()]
    if not parts:
        parts = [
            str(node.get("level1") or "").strip().lower(),
            str(node.get("level2") or "").strip().lower(),
            str(node.get("level3") or "").strip().lower(),
        ]
        parts = [part for part in parts if part]
    return " > ".join(parts)


def _merge_llm_field_candidates(candidates: list[FieldCandidate], llm_result: dict[str, Any]) -> list[FieldCandidate]:
    if not llm_result.get("enabled") or llm_result.get("fallback_used"):
        return candidates
    by_name = {field.name: field for field in candidates}
    selectors = llm_result.get("selectors") if isinstance(llm_result.get("selectors"), dict) else {}
    target_fields = [
        _normalize_field_name(str(item))
        for item in (llm_result.get("target_fields") or [])
        if str(item).strip()
    ]
    for name in target_fields:
        if name in by_name:
            current = by_name[name]
            selector = _safe_selector(selectors.get(name)) or current.selector
            by_name[name] = FieldCandidate(
                name=current.name,
                label=current.label,
                selected=True,
                source="llm" if selector and selector != current.selector else current.source,
                selector=selector,
                api_path=current.api_path,
                confidence=max(current.confidence, 0.65),
                reason=str(llm_result.get("reasoning_summary") or "selected by LLM site analysis")[:240],
            )
        elif name:
            by_name[name] = FieldCandidate(
                name=name,
                label=_field_label(name),
                selected=True,
                source="llm",
                selector=_safe_selector(selectors.get(name)),
                confidence=0.6,
                reason=str(llm_result.get("reasoning_summary") or "suggested by LLM site analysis")[:240],
            )
    for name, selector_value in selectors.items():
        normalized = _normalize_field_name(str(name))
        selector = _safe_selector(selector_value)
        if not normalized or normalized == "item_container" or not selector:
            continue
        if normalized in by_name:
            current = by_name[normalized]
            if not current.selector:
                by_name[normalized] = FieldCandidate(
                    name=current.name,
                    label=current.label,
                    selected=current.selected,
                    source="llm",
                    selector=selector,
                    api_path=current.api_path,
                    confidence=max(current.confidence, 0.6),
                    reason=current.reason or "selector suggested by LLM site analysis",
                )
    ordered = []
    seen = set()
    for field in candidates:
        ordered.append(by_name[field.name])
        seen.add(field.name)
    for name, field in by_name.items():
        if name not in seen:
            ordered.append(field)
    return ordered


def _safe_selector(value: Any) -> str:
    selector = str(value or "").strip()
    if not selector or len(selector) > 300:
        return ""
    if any(ord(ch) < 32 for ch in selector):
        return ""
    return selector


def _dedupe_preserve_order(items: list[str]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for item in items:
        value = item.strip()
        if value and value not in seen:
            output.append(value)
            seen.add(value)
    return output


def discover_catalog_tree(html: str, *, base_url: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html or "", "lxml")
    candidates: list[tuple[str, str]] = []
    scopes = soup.select("nav a[href], header a[href], [role=navigation] a[href], .menu a[href], .nav a[href]")
    if not scopes:
        scopes = soup.select("a[href]")
    for anchor in scopes:
        label = _clean_text(anchor.get_text(" ", strip=True))
        href = str(anchor.get("href") or "").strip()
        if not label or not href:
            continue
        url = urljoin(base_url, href)
        if not _looks_like_catalog_url(url, label):
            continue
        candidates.append((label, url))

    deduped: dict[str, tuple[str, str]] = {}
    for label, url in candidates:
        deduped.setdefault(_url_without_fragment(url), (label, _url_without_fragment(url)))
    nodes = [
        CatalogNode(
            id=_stable_id([label], url),
            label=label,
            url=url,
            path=[label],
            level1=label,
            source="agent",
        ).to_dict()
        for label, url in list(deduped.values())[:200]
    ]
    return nodes


def discover_catalog_from_site_fallbacks(base_url: str, html: str = "") -> list[dict[str, Any]]:
    """Discover catalog-like URLs when the entry page is a JS shell.

    Many ecommerce front pages render navigation client-side, so the initial
    HTML may contain only scripts and a small "enable JavaScript" fallback. This
    helper gathers conservative category candidates from common sitemap paths,
    public JS route strings, and Magento/PWA route hints. It is intentionally
    bounded so analysis stays fast enough for the frontend.
    """
    graphql_nodes = _catalog_nodes_from_graphql_endpoint(base_url)
    if graphql_nodes:
        return graphql_nodes

    candidates: list[tuple[str, str]] = []
    candidates.extend(_catalog_candidates_from_html_assets(base_url, html))
    candidates.extend(_catalog_candidates_from_common_paths(base_url))
    return _catalog_nodes_from_candidates(candidates)


def _catalog_nodes_from_graphql_endpoint(base_url: str) -> list[dict[str, Any]]:
    parsed = urlparse(base_url)
    if not parsed.scheme or not parsed.netloc:
        return []
    endpoint = f"{parsed.scheme}://{parsed.netloc}/graphql"
    query = """
    query getMegaMenu {
      storeConfig {
        base_url
        secure_base_url
        category_url_suffix
      }
      categoryList {
        uid
        name
        include_in_menu
        level
        url_path
        children {
          uid
          name
          include_in_menu
          level
          url_path
          children {
            uid
            name
            include_in_menu
            level
            url_path
            children {
              uid
              name
              include_in_menu
              level
              url_path
            }
          }
        }
      }
    }
    """
    try:
        with httpx.Client(
            follow_redirects=True,
            timeout=httpx.Timeout(10.0, connect=4.0),
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
                "Accept": "application/json",
                "Content-Type": "application/json",
            },
        ) as client:
            response = client.post(endpoint, json={"query": query, "operationName": "getMegaMenu"})
        if response.status_code >= 400:
            return []
        payload = response.json()
    except Exception:
        return []

    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, dict):
        return []
    store_config = data.get("storeConfig") if isinstance(data.get("storeConfig"), dict) else {}
    root = str(store_config.get("secure_base_url") or store_config.get("base_url") or f"{parsed.scheme}://{parsed.netloc}/")
    suffix = str(store_config.get("category_url_suffix") or "")
    category_list = data.get("categoryList") if isinstance(data.get("categoryList"), list) else []
    nodes: list[CatalogNode] = []
    for category in category_list:
        nodes.extend(_catalog_nodes_from_graphql_category(category, root=root, suffix=suffix, path=[]))
    return [node.to_dict() for node in nodes[:200]]


def _catalog_nodes_from_graphql_category(
    category: Any,
    *,
    root: str,
    suffix: str,
    path: list[str],
) -> list[CatalogNode]:
    if not isinstance(category, dict):
        return []
    include = category.get("include_in_menu")
    name = _clean_catalog_label(str(category.get("name") or ""))
    url_path = str(category.get("url_path") or "").strip("/")
    next_path = [*path, name] if name else list(path)
    children_raw = category.get("children") if isinstance(category.get("children"), list) else []
    child_nodes: list[CatalogNode] = []
    for child in children_raw:
        child_nodes.extend(_catalog_nodes_from_graphql_category(child, root=root, suffix=suffix, path=next_path))

    if include is False:
        return child_nodes
    if not name or not url_path:
        return child_nodes

    url = _graphql_category_url(root, url_path, suffix)
    if not _looks_like_catalog_url(url, name):
        return child_nodes
    uid = str(category.get("uid") or "").strip()
    return [
        CatalogNode(
            id=_stable_id(next_path, url),
            label=name,
            url=url,
            path=next_path,
            level1=_level(next_path, 0),
            level2=_level(next_path, 1),
            level3=_level(next_path, 2),
            children=child_nodes,
            source=f"graphql:{uid}" if uid else "graphql",
        )
    ]


def _graphql_category_url(root: str, url_path: str, suffix: str) -> str:
    path = url_path.strip("/")
    if suffix and not path.endswith(suffix):
        path = f"{path}{suffix}"
    return urljoin(root.rstrip("/") + "/", path)


def _catalog_nodes_from_candidates(candidates: list[tuple[str, str]]) -> list[dict[str, Any]]:
    deduped: dict[str, tuple[str, str]] = {}
    for label, url in candidates:
        cleaned_label = _clean_catalog_label(label)
        cleaned_url = _url_without_fragment(url)
        if not cleaned_label or not cleaned_url:
            continue
        if not _looks_like_catalog_url(cleaned_url, cleaned_label):
            continue
        deduped.setdefault(cleaned_url, (cleaned_label, cleaned_url))
    return [
        CatalogNode(
            id=_stable_id([label], url),
            label=label,
            url=url,
            path=[label],
            level1=label,
            source="fallback",
        ).to_dict()
        for label, url in list(deduped.values())[:200]
    ]


def _catalog_candidates_from_html_assets(base_url: str, html: str) -> list[tuple[str, str]]:
    soup = BeautifulSoup(html or "", "lxml")
    candidates: list[tuple[str, str]] = []
    for anchor in soup.select("a[href]"):
        label = _clean_text(anchor.get_text(" ", strip=True)) or _label_from_url(str(anchor.get("href") or ""))
        href = str(anchor.get("href") or "").strip()
        if href:
            candidates.append((label, urljoin(base_url, href)))
    script_urls = [
        urljoin(base_url, str(script.get("src") or ""))
        for script in soup.select("script[src]")
        if script.get("src")
    ]
    for script_url in script_urls[:8]:
        text = _fetch_text_quietly(script_url, max_chars=1_500_000)
        if not text:
            continue
        candidates.extend(_catalog_candidates_from_js_text(base_url, text))
    return candidates


def _catalog_candidates_from_js_text(base_url: str, text: str) -> list[tuple[str, str]]:
    candidates: list[tuple[str, str]] = []
    for match in re.finditer(r"""["'](?P<path>/[A-Za-z0-9][^"'<>\\\s]{1,160})["']""", text or ""):
        path = match.group("path")
        if path.startswith("//") or _looks_like_asset_path(path) or ":" in path:
            continue
        label = _label_from_url(path)
        candidates.append((label, urljoin(base_url, path)))
    return candidates[:500]


def _catalog_candidates_from_common_paths(base_url: str) -> list[tuple[str, str]]:
    candidates: list[tuple[str, str]] = []
    parsed = urlparse(base_url)
    root = f"{parsed.scheme}://{parsed.netloc}/"
    common_paths = [
        "/shop",
        "/sklep",
        "/catalog",
        "/category",
        "/categories",
        "/collections",
        "/produkty",
        "/rowery",
        "/akcesoria",
        "/odziez",
        "/sale",
        "/outlet",
    ]
    for path in common_paths:
        label = _label_from_url(path)
        candidates.append((label, urljoin(root, path.lstrip("/"))))
    for sitemap_url in _candidate_sitemap_urls(root):
        xml = _fetch_text_quietly(sitemap_url, max_chars=500_000)
        if xml:
            candidates.extend(_catalog_candidates_from_sitemap_xml(xml))
    return candidates


def _candidate_sitemap_urls(root: str) -> list[str]:
    return [
        urljoin(root, "sitemap.xml"),
        urljoin(root, "sitemap_index.xml"),
        urljoin(root, "sitemap/sitemap.xml"),
        urljoin(root, "media/sitemap.xml"),
    ]


def _catalog_candidates_from_sitemap_xml(xml: str) -> list[tuple[str, str]]:
    candidates: list[tuple[str, str]] = []
    for loc in re.findall(r"<loc>\s*([^<]+?)\s*</loc>", xml or "", flags=re.I):
        url = loc.strip()
        candidates.append((_label_from_url(url), url))
    return candidates[:500]


def _fetch_text_quietly(url: str, *, max_chars: int) -> str:
    try:
        with httpx.Client(
            follow_redirects=True,
            timeout=httpx.Timeout(12.0, connect=6.0),
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36"},
        ) as client:
            response = client.get(url)
            if response.status_code >= 400:
                return ""
            return response.text[:max_chars]
    except Exception:
        return ""


def _looks_like_asset_path(path: str) -> bool:
    lower = path.lower()
    return any(
        lower.endswith(suffix)
        for suffix in (".js", ".css", ".png", ".jpg", ".jpeg", ".webp", ".svg", ".ico", ".woff", ".woff2", ".map")
    ) or any(token in lower for token in ("/static", "/media/", "/fonts/", "/images/", "/assets/"))


def discover_field_candidates(recon: dict[str, Any], *, field_goal: str = "") -> list[FieldCandidate]:
    dom = recon.get("dom_structure") if isinstance(recon.get("dom_structure"), dict) else {}
    selectors = dom.get("field_selectors") if isinstance(dom.get("field_selectors"), dict) else {}
    target_fields = set(_fields_from_text(field_goal))
    fields: list[FieldCandidate] = []
    for name in DEFAULT_PRODUCT_FIELDS:
        selector = _selector_for_field(name, selectors)
        fields.append(
            FieldCandidate(
                name=name,
                label=_field_label(name),
                selected=not target_fields or name in target_fields,
                source="dom" if selector else "default",
                selector=selector,
                confidence=0.85 if selector else 0.45,
                reason="detected from page structure" if selector else "standard ecommerce field",
            )
        )
    for name in sorted(target_fields - {field.name for field in fields}):
        fields.append(
            FieldCandidate(
                name=name,
                label=_field_label(name),
                selected=True,
                source="user_goal",
                confidence=0.4,
                reason="requested by natural language goal",
            )
        )
    return fields


def resolve_fields(
    available_fields: list[dict[str, Any]],
    *,
    natural_language: str = "",
    requested_fields: list[str] | None = None,
) -> dict[str, Any]:
    selected = set(_fields_from_text(natural_language))
    for item in requested_fields or []:
        selected.add(_normalize_field_name(str(item)))
    if not selected:
        selected = {
            str(item.get("name"))
            for item in available_fields
            if isinstance(item, dict) and item.get("selected", True)
        }
    resolved = []
    missing = []
    known = {str(item.get("name")): item for item in available_fields if isinstance(item, dict)}
    for name in sorted(item for item in selected if item):
        if name in known:
            resolved.append({**known[name], "selected": True})
        else:
            missing.append(name)
            resolved.append(FieldCandidate(
                name=name,
                label=_field_label(name),
                selected=True,
                source="user_goal",
                confidence=0.25,
                reason="not detected yet; requires advisor/browser/API refinement",
            ).to_dict())
    return {
        "schema_version": "field-resolution/v1",
        "selected_fields": [item["name"] for item in resolved],
        "resolved_fields": resolved,
        "missing_fields": missing,
        "needs_refinement": bool(missing),
    }


def build_run_spec(payload: dict[str, Any]) -> CrawlRunSpec:
    export_payload = payload.get("export") if isinstance(payload.get("export"), dict) else {}
    template_data = export_payload.get("template") if isinstance(export_payload.get("template"), dict) else None
    return CrawlRunSpec(
        target_url=str(payload.get("target_url") or ""),
        profile=dict(payload.get("profile") or {}),
        catalog_nodes=list(payload.get("catalog_nodes") or []),
        selected_fields=[str(item) for item in payload.get("selected_fields") or [] if str(item).strip()],
        export=ExportSpec(
            format=str(export_payload.get("format") or "xlsx"),
            output_path=str(export_payload.get("output_path") or ""),
            template_path=str(export_payload.get("template_path") or ""),
            field_mapping={str(k): str(v) for k, v in dict(export_payload.get("field_mapping") or {}).items()},
            template=ExportTemplate.from_dict(template_data) if template_data else ExportTemplate(),
        ),
        run_mode=str(payload.get("run_mode") or "direct"),
        item_workers=max(1, int(payload.get("item_workers") or 4)),
        max_sites=max(1, min(int(payload.get("max_sites") or 1), 5)),
        test_limit=max(1, int(payload.get("test_limit") or 100)),
        runtime_dir=str(payload.get("runtime_dir") or ""),
    )


def profile_from_run_spec(spec: CrawlRunSpec, *, limit: int | None = None) -> SiteProfile:
    profile_data = dict(spec.profile or {})
    run_catalog_nodes = list(spec.catalog_nodes or [])
    selected_nodes = _flatten_catalog_nodes(run_catalog_nodes)
    if not profile_data.get("api_hints") and selected_nodes and not any(node.get("graphql_category_uid") for node in selected_nodes):
        discovered_nodes = discover_catalog_from_site_fallbacks(spec.target_url)
        if discovered_nodes:
            run_catalog_nodes = _enrich_catalog_with_discovered_metadata(run_catalog_nodes, discovered_nodes)
            selected_nodes = _flatten_catalog_nodes(run_catalog_nodes)
            existing_prefs = dict(profile_data.get("crawl_preferences") or {})
            if isinstance(existing_prefs.get("catalog_tree"), list):
                existing_prefs["catalog_tree"] = _enrich_catalog_with_discovered_metadata(
                    list(existing_prefs.get("catalog_tree") or []),
                    discovered_nodes,
                )
            else:
                existing_prefs["catalog_tree"] = run_catalog_nodes
            profile_data["crawl_preferences"] = existing_prefs
    seed_urls = [node["url"] for node in selected_nodes if node.get("url")]
    crawl_preferences = dict(profile_data.get("crawl_preferences") or {})
    if seed_urls:
        crawl_preferences["seed_urls"] = seed_urls
        crawl_preferences.setdefault("seed_kind", _seed_kind_for_urls(seed_urls))
    elif not crawl_preferences.get("seed_urls") and spec.target_url:
        crawl_preferences["seed_urls"] = [spec.target_url]
        crawl_preferences.setdefault("seed_kind", _seed_kind_for_urls([spec.target_url]))
    if limit is not None:
        crawl_preferences["max_items"] = int(limit)
    api_hints = _graphql_products_api_hints(spec.target_url, selected_nodes, limit)
    if api_hints and not profile_data.get("api_hints"):
        profile_data["api_hints"] = api_hints
        crawl_preferences["seed_urls"] = [api_hints["endpoint"]]
        crawl_preferences["seed_kind"] = "api"
        pagination_hints = dict(profile_data.get("pagination_hints") or {})
        pagination_hints.update({
            "type": "page",
            "page_param": "currentPage",
            "start_page": 1,
            "page_size": int(api_hints.get("page_size") or limit or 50),
        })
        profile_data["pagination_hints"] = pagination_hints
    profile_data["crawl_preferences"] = crawl_preferences
    profile_data["access_config"] = _normalize_run_access_config(
        profile_data.get("access_config"),
        profile_data,
        crawl_preferences,
    )
    profile_data["selectors"] = _normalize_run_selectors(profile_data.get("selectors"), spec.selected_fields)
    profile_data["pagination_hints"] = _normalize_run_pagination_hints(
        profile_data.get("pagination_hints"),
        spec.target_url,
    )
    profile_data["training_notes"] = _append_training_notes(
        profile_data.get("training_notes"),
        _run_readiness_notes(profile_data, crawl_preferences),
    )
    if spec.selected_fields:
        profile_data["target_fields"] = list(spec.selected_fields)
    quality = dict(profile_data.get("quality_expectations") or {})
    if spec.selected_fields:
        quality["required_fields"] = list(spec.selected_fields)
    profile_data["quality_expectations"] = quality
    if not profile_data.get("name"):
        profile_data["name"] = urlparse(spec.target_url).netloc or "frontend-run"
    return SiteProfile.from_dict(profile_data)


def _profile_runtime_mode(profile: SiteProfile) -> str:
    access = profile.access_config if isinstance(profile.access_config, dict) else {}
    mode = str(access.get("mode") or access.get("runtime_mode") or "static").strip().lower()
    if mode in {"browser", "playwright"}:
        return "dynamic"
    if mode in {"dynamic", "protected"}:
        return mode
    return "static"


def _graphql_products_api_hints(
    target_url: str,
    selected_nodes: list[dict[str, Any]],
    limit: int | None,
) -> dict[str, Any]:
    category_uid = ""
    category_label = ""
    for node in selected_nodes:
        if not isinstance(node, dict):
            continue
        category_uid = str(node.get("graphql_category_uid") or "").strip()
        category_label = str(node.get("label") or node.get("level3") or node.get("level2") or "").strip()
        if category_uid:
            break
    if not category_uid:
        return {}
    parsed = urlparse(target_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return {}
    endpoint = f"{parsed.scheme}://{parsed.netloc}/graphql"
    page_size = max(1, min(int(limit or 50), 100))
    query = """
    query CLMProducts($categoryUid: String!, $pageSize: Int!, $currentPage: Int!) {
      products(filter: { category_uid: { eq: $categoryUid } }, pageSize: $pageSize, currentPage: $currentPage) {
        total_count
        items {
          name
          sku
          url_key
          url_suffix
          url_path
          description { html }
          short_description { html }
          image { url label }
          small_image { url label }
          thumbnail { url label }
          price_range {
            maximum_price {
              regular_price { value currency }
              final_price { value currency }
            }
            minimum_price {
              regular_price { value currency }
              final_price { value currency }
            }
          }
        }
      }
    }
    """.strip()
    return {
        "endpoint": endpoint,
        "method": "POST",
        "kind": "api",
        "format": "graphql_products",
        "items_path": "data.products.items",
        "total_path": "data.products.total_count",
        "page_size": page_size,
        "post_json": {
            "query": query,
            "variables": {
                "categoryUid": category_uid,
                "pageSize": page_size,
                "currentPage": 1,
            },
        },
        "field_mapping": {
            "title": "name",
            "canonical_url": "url_key",
            "highest_price": [
                "price_range.maximum_price.regular_price.value",
                "price_range.maximum_price.final_price.value",
            ],
            "currency": [
                "price_range.maximum_price.regular_price.currency",
                "price_range.maximum_price.final_price.currency",
            ],
            "description": ["description.html", "short_description.html"],
            "image_urls": ["image.url", "small_image.url", "thumbnail.url"],
        },
        "category_uid": category_uid,
        "category": category_label,
    }


def _seed_kind_for_urls(urls: list[str]) -> str:
    if urls and all(_looks_like_product_detail_url(url) for url in urls):
        return "detail"
    return "list"


def _looks_like_product_detail_url(url: str) -> bool:
    path = urlparse(str(url or "")).path.lower()
    tokens = set(_catalog_match_tokens(path, ""))
    detail_tokens = {
        "product", "products", "produkt", "produkty", "p", "pd", "item", "sku",
        "towar", "artykul", "detail", "details",
    }
    if tokens & detail_tokens:
        return True
    if re.search(r"(^|/)(p|pd|sku)[-/]?[a-z0-9]{4,}($|/|\\.)", path):
        return True
    if path.endswith(".html") and re.search(r"[a-z]+-[a-z0-9-]*\\d{2,}", path):
        return True
    return False


def _normalize_run_selectors(value: Any, selected_fields: list[str]) -> dict[str, Any]:
    selectors = dict(value) if isinstance(value, dict) else {}
    normalized: dict[str, Any] = dict(selectors)
    flat_fields = {
        str(key): spec
        for key, spec in selectors.items()
        if key not in {"list", "category", "detail"} and isinstance(spec, (str, dict))
    }
    if flat_fields and not isinstance(normalized.get("detail"), dict):
        normalized["detail"] = dict(flat_fields)
    detail_selectors = dict(normalized.get("detail") or {}) if isinstance(normalized.get("detail"), dict) else {}
    for field in selected_fields or []:
        fallback = _fallback_detail_selector(field)
        if fallback and field not in detail_selectors:
            detail_selectors[field] = fallback
    if detail_selectors:
        normalized["detail"] = detail_selectors
    normalized.setdefault("list", {})
    normalized.setdefault("category", dict(normalized.get("list") or {}))
    return normalized


def _fallback_detail_selector(field: str) -> Any:
    return {
        "title": {
            "selector_type": "xpath",
            "selector": "string((//h1 | //*[@itemprop='name'] | //meta[@property='og:title']/@content | //title)[1])",
            "many": False,
        },
        "highest_price": {
            "selector_type": "xpath",
            "selector": "string((//meta[@property='product:price:amount']/@content | //*[@itemprop='price']/@content | //*[@itemprop='price'])[1])",
            "many": False,
        },
        "price": {
            "selector_type": "xpath",
            "selector": "string((//meta[@property='product:price:amount']/@content | //*[@itemprop='price']/@content | //*[@itemprop='price'])[1])",
            "many": False,
        },
        "description": {
            "selector_type": "xpath",
            "selector": "string((//*[@itemprop='description'] | //meta[@name='description']/@content |//*[contains(@class,'description')])[1])",
            "many": False,
        },
        "image_urls": {
            "selector_type": "xpath",
            "selector": "//meta[@property='og:image']/@content | //*[@itemprop='image']/@src | //img/@src",
            "many": True,
        },
        "colors": "[class*='color'], [data-color], [aria-label*='color'], [aria-label*='kolor']",
        "sizes": "[class*='size'], [data-size], [aria-label*='size'], [aria-label*='rozmiar']",
    }.get(str(field), "")


def _normalize_run_pagination_hints(value: Any, target_url: str) -> dict[str, Any]:
    hints = dict(value) if isinstance(value, dict) else {}
    link_hints = hints.get("link_discovery") if isinstance(hints.get("link_discovery"), dict) else {}
    if not link_hints:
        parsed = urlparse(target_url)
        domain = parsed.netloc.lower()
        link_hints = {
            "allow_domains": [domain] if domain else [],
            "allow": [
                r"/product/",
                r"/products/",
                r"/produkt/",
                r"/produkty/",
                r"/p/",
                r"\.html(?:\?|$)",
            ],
            "deny": [
                r"/cart",
                r"/checkout",
                r"/account",
                r"/login",
                r"/privacy",
                r"/terms",
                r"/blog",
                r"/contact",
            ],
            "classify": {
                "detail": r"(/product/|/products/|/produkt/|/produkty/|/p/|/[a-z0-9-]*-[a-z0-9-]*\d{2,}[^/]*\.html(?:\?|$))",
                "list": r"(/category/|/categories/|/collections/|/catalog|/shop|/sklep|/rowery|/akcesoria|/czesci|/marki)",
            },
            "default_kind": "list",
            "max_links": 300,
        }
    hints["type"] = str(hints.get("type") or "dom_links")
    hints["link_discovery"] = link_hints
    return hints


def _normalize_run_access_config(
    value: Any,
    profile_data: dict[str, Any],
    crawl_preferences: dict[str, Any],
) -> dict[str, Any]:
    access = dict(value) if isinstance(value, dict) else {}
    mode = str(access.get("mode") or access.get("runtime_mode") or "").strip().lower()
    if not mode and _profile_needs_dynamic_runtime(profile_data, crawl_preferences):
        mode = "dynamic"
    if mode in {"browser", "playwright"}:
        mode = "dynamic"
    if mode in {"dynamic", "protected"}:
        access["mode"] = mode
        access.setdefault("wait_until", "domcontentloaded")
        browser_config = dict(access.get("browser_config") or {})
        browser_config.setdefault("render_time_ms", 5000)
        browser_config.setdefault("auto_accept_cookies", True)
        browser_config.setdefault("capture_api", True)
        access["browser_config"] = browser_config
    return access


def _profile_needs_dynamic_runtime(profile_data: dict[str, Any], crawl_preferences: dict[str, Any]) -> bool:
    notes = " ".join(str(item).lower() for item in profile_data.get("training_notes") or [])
    if any(token in notes for token in ("js shell", "pwa", "spa", "dynamic")):
        return True
    rendering = str(profile_data.get("rendering") or profile_data.get("recon_summary", {}).get("rendering") or "").lower()
    if rendering in {"spa", "pwa", "client"}:
        return True
    seed_urls = [str(item).lower() for item in crawl_preferences.get("seed_urls") or []]
    return bool(seed_urls and all(url.endswith(".html") for url in seed_urls) and len(seed_urls) >= 10)


def _append_training_notes(existing: Any, additions: list[str]) -> list[str]:
    notes = [str(item) for item in (existing or []) if str(item).strip()] if isinstance(existing, list) else []
    for note in additions:
        if note and note not in notes:
            notes.append(note)
    return notes


def _run_readiness_notes(profile_data: dict[str, Any], crawl_preferences: dict[str, Any]) -> list[str]:
    notes: list[str] = []
    seed_urls = [str(item) for item in crawl_preferences.get("seed_urls") or [] if str(item).strip()]
    if not seed_urls:
        notes.append("run_readiness:no seed URLs; runner cannot crawl")
    selectors = profile_data.get("selectors") if isinstance(profile_data.get("selectors"), dict) else {}
    detail_selectors = selectors.get("detail") if isinstance(selectors.get("detail"), dict) else selectors
    if not isinstance(detail_selectors, dict) or not detail_selectors.get("title"):
        notes.append("run_readiness:no detail title selector; generic metadata fallback injected")
    if crawl_preferences.get("seed_kind") == "list":
        notes.append("run_readiness:list seeds require link discovery before product extraction")
    return notes


def build_test_run_payload(spec: CrawlRunSpec) -> dict[str, Any]:
    return {
        "profile": profile_from_run_spec(spec, limit=spec.test_limit).to_dict(),
        "run_id": f"test-{uuid.uuid4().hex[:8]}",
        "batch_size": min(max(spec.item_workers * 2, 10), 100),
        "max_batches": max(1, (spec.test_limit // max(spec.item_workers * 2, 10)) + 1),
        "item_workers": spec.item_workers,
        "runtime_dir": spec.runtime_dir,
    }


def build_full_run_payload(spec: CrawlRunSpec) -> dict[str, Any]:
    return {
        "profile": profile_from_run_spec(spec, limit=None).to_dict(),
        "run_id": f"full-{uuid.uuid4().hex[:8]}",
        "batch_size": min(max(spec.item_workers * 4, 20), 200),
        "max_batches": 0,
        "item_workers": spec.item_workers,
        "runtime_dir": spec.runtime_dir,
    }


def summarize_run_progress(job: dict[str, Any]) -> dict[str, Any]:
    profile_run = job.get("profile_run") if isinstance(job.get("profile_run"), dict) else {}
    summary = profile_run.get("runner_summary") if isinstance(profile_run.get("runner_summary"), dict) else {}
    frontier = profile_run.get("frontier_stats") if isinstance(profile_run.get("frontier_stats"), dict) else {}
    product_stats = profile_run.get("product_stats") if isinstance(profile_run.get("product_stats"), dict) else {}
    claimed = int(summary.get("claimed") or 0)
    saved = int(summary.get("records_saved") or product_stats.get("total") or job.get("item_count") or 0)
    failed = int(summary.get("failed") or frontier.get("failed") or 0)
    failure_buckets = summary.get("failure_buckets") if isinstance(summary.get("failure_buckets"), dict) else {}
    queued = int(frontier.get("queued") or 0)
    done = int(frontier.get("done") or summary.get("succeeded") or 0)
    total_known = max(done + queued + failed, claimed, saved)
    completion = round(done / total_known, 4) if total_known else 0.0
    quality = profile_run.get("quality_summary", {})
    status = job.get("status", "unknown")

    # Current stage: derive from status and progress
    current_stage = _derive_current_stage(status, queued, done, failed, total_known)

    # Last error snippet: most recent failure message, truncated
    last_error = _extract_last_error_snippet(job, profile_run)

    # Progress summary: human-readable one-liner
    progress_summary = _build_progress_summary(status, saved, done, failed, queued, completion)

    # Quality indicator: pass/warn/fail based on field coverage
    quality_indicator = _derive_quality_indicator(quality, saved, failed)

    return {
        "status": status,
        "records_saved": saved,
        "claimed": claimed,
        "failed": failed,
        "queued": queued,
        "done": done,
        "completion": completion,
        "estimated_remaining_seconds": None,
        "quality": quality,
        "failure_buckets": dict(failure_buckets),
        "current_stage": current_stage,
        "last_error": last_error,
        "progress_summary": progress_summary,
        "quality_indicator": quality_indicator,
    }


def build_run_evidence_pack(job: dict[str, Any]) -> dict[str, Any]:
    """Build a compact, actionable evidence pack for managed decisions."""
    profile_run = job.get("profile_run") if isinstance(job.get("profile_run"), dict) else {}
    run_spec = job.get("product_run_spec") if isinstance(job.get("product_run_spec"), dict) else {}
    profile = _extract_profile_from_job(job, profile_run)
    target_url = str(run_spec.get("target_url") or job.get("target_url") or "")
    progress = summarize_run_progress(job)
    quality = progress.get("quality") if isinstance(progress.get("quality"), dict) else {}
    diagnostics = job.get("diagnostics") if isinstance(job.get("diagnostics"), dict) else {}
    supervision = job.get("supervision") if isinstance(job.get("supervision"), dict) else {}
    runner = profile_run.get("runner_summary") if isinstance(profile_run.get("runner_summary"), dict) else {}
    frontier = profile_run.get("frontier_stats") if isinstance(profile_run.get("frontier_stats"), dict) else {}
    product_stats = profile_run.get("product_stats") if isinstance(profile_run.get("product_stats"), dict) else {}

    return {
        "schema_version": "run-evidence-pack/v1",
        "task": {
            "task_id": job.get("task_id", ""),
            "kind": job.get("kind", ""),
            "status": job.get("status", ""),
            "target_url": run_spec.get("target_url") or job.get("target_url") or "",
            "run_id": job.get("run_id", ""),
        },
        "progress": progress,
        "run_spec_summary": {
            "selected_fields": list(run_spec.get("selected_fields") or []),
            "catalog_node_count": len(run_spec.get("catalog_nodes") or []) if isinstance(run_spec.get("catalog_nodes"), list) else 0,
            "item_workers": run_spec.get("item_workers"),
            "test_limit": run_spec.get("test_limit"),
            "runtime_dir": run_spec.get("runtime_dir", ""),
            "export": run_spec.get("export") if isinstance(run_spec.get("export"), dict) else {},
        },
        "profile_summary": _profile_evidence_summary(profile),
        "quality_gaps": _quality_gap_evidence(quality, run_spec),
        "access_evidence": build_access_evidence_snapshot(job),
        "failure_evidence": {
            "failure_buckets": dict(progress.get("failure_buckets") or {}),
            "last_error": progress.get("last_error", ""),
            "recent_failures": _recent_failure_evidence(profile_run),
            "runner": {
                "claimed": runner.get("claimed", 0),
                "succeeded": runner.get("succeeded", 0),
                "failed": runner.get("failed", 0),
                "records_saved": runner.get("records_saved", product_stats.get("total", 0)),
                "checkpoint_errors": runner.get("checkpoint_errors", 0),
            },
            "frontier": dict(frontier),
        },
        "diagnostics": _compact_dict({
            "runner_summary": diagnostics.get("runner_summary") if isinstance(diagnostics, dict) else {},
            "backpressure": diagnostics.get("backpressure") if isinstance(diagnostics, dict) else {},
            "supervision": supervision,
            "coverage_report": diagnostics.get("coverage_report") if isinstance(diagnostics, dict) else {},
            "throughput": diagnostics.get("throughput") if isinstance(diagnostics, dict) else {},
        }, max_items=30),
        "managed_history": {
            "ai_decision_count": len(job.get("ai_decisions") or []) if isinstance(job.get("ai_decisions"), list) else 0,
            "llm_trace_count": len(job.get("llm_traces") or []) if isinstance(job.get("llm_traces"), list) else 0,
            "managed_action_count": len(job.get("managed_actions") or []) if isinstance(job.get("managed_actions"), list) else 0,
            "managed_step_count": len(job.get("managed_steps") or []) if isinstance(job.get("managed_steps"), list) else 0,
            "managed_control_loop_count": len(job.get("managed_control_loops") or []) if isinstance(job.get("managed_control_loops"), list) else 0,
            "latest_llm_errors": _latest_llm_error_evidence(job),
        },
        "recommended_focus": _recommended_focus(progress, quality, supervision),
    }


def build_access_evidence_snapshot(job: dict[str, Any]) -> dict[str, Any]:
    """Extract compact access/runtime evidence already produced by the backend.

    This is intentionally a sampler, not another crawler. It turns profile-run
    diagnostics, runtime events, failure buckets, browser/XHR artifacts, and
    recent failures into a small packet that an LLM can reason over without
    reading full raw job state.
    """
    profile_run = job.get("profile_run") if isinstance(job.get("profile_run"), dict) else {}
    run_spec = job.get("product_run_spec") if isinstance(job.get("product_run_spec"), dict) else {}
    profile = _extract_profile_from_job(job, profile_run)
    target_url = str(run_spec.get("target_url") or job.get("target_url") or "")
    progress = summarize_run_progress(job)
    recent_failures = _recent_failure_evidence(profile_run)
    profile_summary = _profile_evidence_summary(profile)
    runtime_events = _runtime_event_samples(profile_run)
    xhr_samples = _xhr_evidence_samples(profile_run)
    artifact_samples = _artifact_evidence_samples(profile_run)
    probe_snapshot = _latest_access_probe_snapshot(job)
    probe_runtime_events = _runtime_event_samples_from_probe(probe_snapshot)
    probe_xhr_samples = _xhr_evidence_samples_from_probe(probe_snapshot)
    probe_artifact_samples = _artifact_evidence_samples_from_probe(probe_snapshot)
    if probe_runtime_events:
        runtime_events = _merge_samples(runtime_events, probe_runtime_events, key_fields=("type", "url", "message"))[:30]
    if probe_xhr_samples:
        xhr_samples = _merge_samples(xhr_samples, probe_xhr_samples, key_fields=("method", "url", "status"))[:30]
    if probe_artifact_samples:
        artifact_samples = _merge_samples(artifact_samples, probe_artifact_samples, key_fields=("kind", "path", "url"))[:30]
    bucket_names = sorted(str(key) for key in (progress.get("failure_buckets") or {}).keys())
    challenge_hits = [
        item for item in recent_failures
        if _text_has_challenge_signals(" ".join(str(value) for value in item.values()))
    ]
    if probe_snapshot and bool(probe_snapshot.get("summary", {}).get("challenge_like")):
        challenge_hits.append({
            "url": str(probe_snapshot.get("final_url") or probe_snapshot.get("target_url") or ""),
            "bucket": "challenge_like",
            "error": str(probe_snapshot.get("error") or probe_snapshot.get("summary", {}).get("framework") or "")[:300],
        })
    if any(name in {"challenge_like", "captcha", "managed_challenge", "http_blocked"} for name in bucket_names):
        challenge_hits.append({
            "url": target_url,
            "bucket": "challenge_like",
            "error": "challenge-like failure bucket present",
        })
    challenge_like = bool(challenge_hits) or any(
        name in {"challenge_like", "captcha", "managed_challenge", "http_blocked"}
        for name in bucket_names
    )
    recommended_runtime = _recommended_access_runtime(
        challenge_like=challenge_like,
        xhr_samples=xhr_samples,
        profile_summary=profile_summary,
        records_saved=int(progress.get("records_saved") or 0),
    )
    missing_evidence = []
    if not runtime_events:
        missing_evidence.append("runtime_events")
    if not xhr_samples:
        missing_evidence.append("xhr_or_api_samples")
    if not artifact_samples:
        missing_evidence.append("browser_artifacts")
    if int(progress.get("records_saved") or 0) == 0 and not recent_failures:
        missing_evidence.append("recent_failures")
    return {
        "schema_version": "access-evidence/v1",
        "target_url": target_url,
        "status": job.get("status", ""),
        "summary": {
            "challenge_like": challenge_like,
            "failure_buckets": dict(progress.get("failure_buckets") or {}),
            "records_saved": int(progress.get("records_saved") or 0),
            "failed": int(progress.get("failed") or 0),
            "access_mode": profile_summary.get("access_mode", ""),
            "recommended_runtime": recommended_runtime,
            "missing_evidence": missing_evidence,
        },
        "profile": profile_summary,
        "recent_failures": recent_failures[:5],
        "challenge_evidence": challenge_hits[:5],
        "runtime_events": runtime_events[:12],
        "xhr_samples": xhr_samples[:10],
        "artifact_samples": artifact_samples[:10],
        "probe_snapshot": probe_snapshot,
        "decision_hints": _access_decision_hints(
            challenge_like=challenge_like,
            xhr_samples=xhr_samples,
            artifact_samples=artifact_samples,
            progress=progress,
            profile_summary=profile_summary,
        ),
    }


def _derive_current_stage(status: str, queued: int, done: int, failed: int, total_known: int) -> str:
    if status in ("failed", "cancelled"):
        return "stopped"
    if status == "completed":
        return "finished"
    if status == "running":
        if queued > 0 and done == 0:
            return "crawling"
        if done > 0 and queued > 0:
            return "crawling"
        if done > 0 and queued == 0:
            return "finishing"
        return "starting"
    return status or "unknown"


def _extract_last_error_snippet(job: dict[str, Any], profile_run: dict[str, Any]) -> str:
    # Check job-level error first
    job_error = str(job.get("error") or "").strip()
    if job_error:
        return job_error[:200]

    # Check profile_run failures
    failures = profile_run.get("failures") or []
    if isinstance(failures, list) and failures:
        last = failures[-1]
        if isinstance(last, dict):
            return str(last.get("error") or last.get("message") or "")[:200]
        return str(last)[:200]

    # Check runner_summary errors
    runner_summary = profile_run.get("runner_summary") if isinstance(profile_run.get("runner_summary"), dict) else {}
    runner_error = str(runner_summary.get("last_error") or "").strip()
    if runner_error:
        return runner_error[:200]

    return ""


def _build_progress_summary(status: str, saved: int, done: int, failed: int, queued: int, completion: float) -> str:
    if status == "completed":
        return f"Done — {saved} records saved"
    if status == "failed":
        return f"Failed after {saved} records ({failed} errors)"
    if status == "running":
        pct = f"{completion * 100:.0f}%"
        parts = [f"{saved} saved"]
        if failed:
            parts.append(f"{failed} failed")
        if queued:
            parts.append(f"{queued} queued")
        return f"Running ({pct}) — {', '.join(parts)}"
    return f"{status} — {saved} records"


def _derive_quality_indicator(quality: dict[str, Any], saved: int, failed: int) -> str:
    if not isinstance(quality, dict):
        quality = {}

    # If we have explicit field coverage, use it
    field_coverage = quality.get("field_coverage")
    if isinstance(field_coverage, (int, float)):
        if field_coverage >= 0.9:
            return "pass"
        if field_coverage >= 0.6:
            return "warn"
        return "fail"

    # Fall back to success rate
    total = saved + failed
    if total == 0:
        return "unknown"
    success_rate = saved / total
    if success_rate >= 0.9:
        return "pass"
    if success_rate >= 0.6:
        return "warn"
    return "fail"


def _extract_profile_from_job(job: dict[str, Any], profile_run: dict[str, Any]) -> dict[str, Any]:
    checkpoint = profile_run.get("checkpoint_latest") if isinstance(profile_run.get("checkpoint_latest"), dict) else {}
    metadata = checkpoint.get("metadata") if isinstance(checkpoint.get("metadata"), dict) else {}
    if isinstance(metadata.get("profile"), dict):
        return metadata["profile"]
    run_spec = job.get("product_run_spec") if isinstance(job.get("product_run_spec"), dict) else {}
    if isinstance(run_spec.get("profile"), dict):
        return run_spec["profile"]
    report = profile_run.get("report") if isinstance(profile_run.get("report"), dict) else {}
    if isinstance(report.get("profile"), dict):
        return report["profile"]
    return {}


def _profile_evidence_summary(profile: dict[str, Any]) -> dict[str, Any]:
    prefs = profile.get("crawl_preferences") if isinstance(profile.get("crawl_preferences"), dict) else {}
    selectors = profile.get("selectors") if isinstance(profile.get("selectors"), dict) else {}
    detail = selectors.get("detail") if isinstance(selectors.get("detail"), dict) else selectors
    access = profile.get("access_config") if isinstance(profile.get("access_config"), dict) else {}
    browser = access.get("browser_config") if isinstance(access.get("browser_config"), dict) else {}
    seed_urls = prefs.get("seed_urls") if isinstance(prefs.get("seed_urls"), list) else []
    return {
        "name": profile.get("name", ""),
        "seed_kind": prefs.get("seed_kind", ""),
        "seed_url_count": len(seed_urls),
        "seed_url_samples": [str(item) for item in seed_urls[:5]],
        "target_fields": list(profile.get("target_fields") or []),
        "selector_keys": sorted(str(key) for key in detail.keys())[:50] if isinstance(detail, dict) else [],
        "access_mode": access.get("mode", ""),
        "wait_until": access.get("wait_until", ""),
        "browser": {
            "capture_api": bool(browser.get("capture_api")),
            "persistent_context": bool(browser.get("persistent_context")),
            "pool_enabled": bool(browser.get("pool_enabled")),
            "max_wait_ms": browser.get("max_wait_ms"),
        },
    }


def _quality_gap_evidence(quality: dict[str, Any], run_spec: dict[str, Any]) -> list[dict[str, Any]]:
    gaps: list[dict[str, Any]] = []
    selected = [str(item) for item in run_spec.get("selected_fields") or [] if str(item).strip()]
    completeness = quality.get("field_completeness") if isinstance(quality.get("field_completeness"), dict) else {}
    for field in selected:
        value = completeness.get(field)
        if isinstance(value, (int, float)) and value < 0.8:
            gaps.append({"field": field, "coverage": round(float(value), 4), "kind": "field_coverage"})
        elif field and field not in completeness and quality:
            gaps.append({"field": field, "coverage": None, "kind": "missing_coverage_metric"})
    if int(quality.get("total_records") or 0) == 0:
        gaps.append({"kind": "zero_records", "field": "", "coverage": 0.0})
    return gaps[:50]


def _recent_failure_evidence(profile_run: dict[str, Any]) -> list[dict[str, Any]]:
    failures = profile_run.get("failures") if isinstance(profile_run.get("failures"), list) else []
    out: list[dict[str, Any]] = []
    for item in failures[-10:]:
        if not isinstance(item, dict):
            out.append({"error": str(item)[:300]})
            continue
        out.append({
            "url": str(item.get("url") or item.get("target_url") or "")[:500],
            "bucket": str(item.get("bucket") or item.get("failure_bucket") or "")[:80],
            "error": str(item.get("error") or item.get("message") or "")[:300],
            "status": item.get("status") or item.get("status_code"),
        })
    return out


def _latest_llm_error_evidence(job: dict[str, Any]) -> list[dict[str, Any]]:
    traces = job.get("llm_traces") if isinstance(job.get("llm_traces"), list) else []
    errors: list[dict[str, Any]] = []
    for trace in traces[-10:]:
        if not isinstance(trace, dict) or not trace.get("error"):
            continue
        errors.append({
            "stage": trace.get("stage", ""),
            "model": trace.get("model", ""),
            "duration_ms": trace.get("duration_ms"),
            "error": str(trace.get("error") or "")[:300],
        })
    return errors


def _runtime_event_samples(profile_run: dict[str, Any]) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    for event in _walk_dict_values(profile_run, keys={"runtime_events", "events", "request_events"}):
        if isinstance(event, list):
            for item in event:
                sample = _runtime_event_sample(item)
                if sample:
                    samples.append(sample)
        else:
            sample = _runtime_event_sample(event)
            if sample:
                samples.append(sample)
    return _dedupe_dict_samples(samples, key_fields=("type", "url", "message"))[:30]


def _runtime_event_sample(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        return {}
    event_type = str(value.get("type") or value.get("event") or value.get("name") or "")[:120]
    message = str(value.get("message") or value.get("error") or value.get("reason") or "")[:300]
    url = str(value.get("url") or value.get("target_url") or value.get("final_url") or "")[:500]
    status = value.get("status") or value.get("status_code")
    bucket = str(value.get("bucket") or value.get("failure_bucket") or value.get("classification") or "")[:120]
    if not any([event_type, message, url, status, bucket]):
        return {}
    return {
        "type": event_type,
        "url": url,
        "status": status,
        "bucket": bucket,
        "message": message,
    }


def _xhr_evidence_samples(profile_run: dict[str, Any]) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    keys = {"captured_xhr", "xhr", "xhr_samples", "api_candidates", "network_observation", "browser_interception"}
    for value in _walk_dict_values(profile_run, keys=keys):
        for candidate in _flatten_listish(value):
            sample = _xhr_evidence_sample(candidate)
            if sample:
                samples.append(sample)
    return _dedupe_dict_samples(samples, key_fields=("method", "url", "status"))[:30]


def _xhr_evidence_sample(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        return {}
    url = str(value.get("url") or value.get("request_url") or value.get("endpoint") or "")[:500]
    method = str(value.get("method") or value.get("request_method") or "GET").upper()[:12]
    status = value.get("status") or value.get("status_code")
    content_type = str(value.get("content_type") or value.get("mime_type") or value.get("resource_type") or "")[:120]
    kind = str(value.get("kind") or value.get("type") or value.get("candidate_type") or "")[:120]
    score = value.get("score")
    preview = value.get("preview") or value.get("body_preview") or value.get("json_preview")
    post_data = value.get("post_data") or value.get("post_data_preview") or value.get("request_body")
    if isinstance(preview, (dict, list)):
        preview_text = json.dumps(preview, ensure_ascii=False, default=str)
    else:
        preview_text = str(preview or "")
    if isinstance(post_data, (dict, list)):
        post_data_text = json.dumps(post_data, ensure_ascii=False, default=str)
    else:
        post_data_text = str(post_data or "")
    if not any([url, status, content_type, kind, preview_text]):
        return {}
    sample = {
        "method": method,
        "url": url,
        "status": status,
        "content_type": content_type,
        "kind": kind,
        "score": score,
        "preview": _redact_sensitive_text(preview_text[:500]),
    }
    if post_data_text.strip():
        sample["post_data_preview"] = _redact_sensitive_text(post_data_text[:2000])
    if isinstance(value.get("request_headers"), dict):
        sample["request_headers"] = _safe_xhr_headers(value.get("request_headers") or {})
    return sample


def _safe_xhr_headers(headers: dict[str, Any]) -> dict[str, str]:
    allowed = {
        "accept", "accept-language", "content-type", "origin", "referer",
        "x-requested-with", "x-csrf-token", "x-xsrf-token", "x-magento-cache-id",
        "x-store", "store",
    }
    output: dict[str, str] = {}
    for key, value in dict(headers or {}).items():
        name = str(key).strip()
        lowered = name.lower()
        if lowered in allowed or lowered.startswith("x-"):
            text = str(value or "").strip()
            if text and len(text) <= 1000 and "\x00" not in text:
                output[name] = text
    return output


def _artifact_evidence_samples(profile_run: dict[str, Any]) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    for value in _walk_dict_values(profile_run, keys={"artifacts", "runtime_artifacts", "artifact_manifest"}):
        for artifact in _flatten_listish(value):
            sample = _artifact_evidence_sample(artifact)
            if sample:
                samples.append(sample)
    return _dedupe_dict_samples(samples, key_fields=("kind", "path", "url"))[:30]


def _artifact_evidence_sample(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        return {}
    kind = str(value.get("kind") or value.get("type") or value.get("artifact_type") or "")[:80]
    path = str(value.get("path") or value.get("file_path") or "")[:500]
    url = str(value.get("url") or value.get("source_url") or "")[:500]
    summary = str(value.get("summary") or value.get("message") or value.get("description") or "")[:300]
    if not any([kind, path, url, summary]):
        return {}
    return {
        "kind": kind,
        "path": path,
        "url": url,
        "summary": summary,
    }


def _latest_access_probe_snapshot(job: dict[str, Any]) -> dict[str, Any]:
    latest = job.get("latest_access_probe") if isinstance(job.get("latest_access_probe"), dict) else {}
    for direct_key in ("probe_snapshot", "snapshot"):
        direct_snapshot = latest.get(direct_key) if isinstance(latest.get(direct_key), dict) else {}
        if direct_snapshot.get("schema_version") == "access-probe/v1":
            return direct_snapshot
    for key in ("managed_steps", "managed_actions", "access_probes", "access_probe_history"):
        values = job.get(key)
        if not isinstance(values, list):
            continue
        for item in reversed(values):
            if not isinstance(item, dict):
                continue
            result = item.get("result") if isinstance(item.get("result"), dict) else item
            for direct_key in ("probe_snapshot", "snapshot"):
                direct_snapshot = result.get(direct_key) if isinstance(result.get(direct_key), dict) else {}
                if direct_snapshot.get("schema_version") == "access-probe/v1":
                    return direct_snapshot
            evidence = result.get("evidence") if isinstance(result.get("evidence"), dict) else {}
            snapshot = evidence.get("snapshot") if isinstance(evidence.get("snapshot"), dict) else {}
            if snapshot.get("schema_version") == "access-probe/v1":
                return snapshot
    return {}


def _runtime_event_samples_from_probe(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    events = snapshot.get("runtime_events") if isinstance(snapshot.get("runtime_events"), list) else []
    return [_runtime_event_sample(item) for item in events if isinstance(item, dict) and _runtime_event_sample(item)]


def _xhr_evidence_samples_from_probe(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    samples = snapshot.get("xhr_samples") if isinstance(snapshot.get("xhr_samples"), list) else []
    return [_xhr_evidence_sample(item) for item in samples if isinstance(item, dict) and _xhr_evidence_sample(item)]


def _artifact_evidence_samples_from_probe(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    samples = snapshot.get("artifact_samples") if isinstance(snapshot.get("artifact_samples"), list) else []
    return [_artifact_evidence_sample(item) for item in samples if isinstance(item, dict) and _artifact_evidence_sample(item)]


def _merge_samples(samples_a: list[dict[str, Any]], samples_b: list[dict[str, Any]], *, key_fields: tuple[str, ...]) -> list[dict[str, Any]]:
    return _dedupe_dict_samples([*samples_a, *samples_b], key_fields=key_fields)


def _access_decision_hints(
    *,
    challenge_like: bool,
    xhr_samples: list[dict[str, Any]],
    artifact_samples: list[dict[str, Any]],
    progress: dict[str, Any],
    profile_summary: dict[str, Any],
) -> list[str]:
    hints: list[str] = []
    if challenge_like:
        hints.append("use_protected_browser_profile")
        hints.append("persist_browser_session")
        hints.append("lower_concurrency_and_prepare_proxy_rotation")
    if xhr_samples:
        hints.append("prefer_api_or_xhr_replay_if_product_payload_is_visible")
    if any("screenshot" in str(item.get("kind") or "").lower() for item in artifact_samples):
        hints.append("use_screenshot_for_visual_or_challenge_review")
    if int(progress.get("records_saved") or 0) == 0:
        hints.append("collect_small_browser_sample_before_bulk_rerun")
    browser = profile_summary.get("browser") if isinstance(profile_summary.get("browser"), dict) else {}
    if not browser.get("capture_api"):
        hints.append("enable_api_capture")
    return list(dict.fromkeys(hints))[:10]


def _recommended_access_runtime(
    *,
    challenge_like: bool,
    xhr_samples: list[dict[str, Any]],
    profile_summary: dict[str, Any],
    records_saved: int,
) -> str:
    if challenge_like:
        return "protected_browser"
    if xhr_samples and records_saved == 0:
        return "api_replay_or_browser_network"
    mode = str(profile_summary.get("access_mode") or "").lower()
    if mode in {"dynamic", "browser", "protected"}:
        return mode
    if records_saved == 0:
        return "dynamic_browser_probe"
    return mode or "static_or_profile_default"


def _walk_dict_values(value: Any, *, keys: set[str]) -> list[Any]:
    found: list[Any] = []
    if isinstance(value, dict):
        for key, item in value.items():
            if str(key) in keys:
                found.append(item)
            if isinstance(item, (dict, list)):
                found.extend(_walk_dict_values(item, keys=keys))
    elif isinstance(value, list):
        for item in value[:200]:
            if isinstance(item, (dict, list)):
                found.extend(_walk_dict_values(item, keys=keys))
    return found


def _flatten_listish(value: Any) -> list[Any]:
    if isinstance(value, list):
        out: list[Any] = []
        for item in value[:200]:
            out.extend(_flatten_listish(item) if isinstance(item, list) else [item])
        return out
    if isinstance(value, dict):
        for key in ("items", "candidates", "responses", "requests", "entries", "artifacts"):
            if isinstance(value.get(key), list):
                return _flatten_listish(value[key])
        return [value]
    return []


def _dedupe_dict_samples(samples: list[dict[str, Any]], *, key_fields: tuple[str, ...]) -> list[dict[str, Any]]:
    seen: set[tuple[str, ...]] = set()
    out: list[dict[str, Any]] = []
    for sample in samples:
        key = tuple(str(sample.get(field) or "") for field in key_fields)
        if key in seen:
            continue
        seen.add(key)
        out.append(sample)
    return out


def _text_has_challenge_signals(text: str) -> bool:
    lowered = text.lower()
    return any(token in lowered for token in (
        "captcha",
        "recaptcha",
        "challenge",
        "cloudflare",
        "datadome",
        "perimeterx",
        "blocked",
        "403",
        "429",
    ))


def _redact_sensitive_text(text: str) -> str:
    text = re.sub(r"(?i)(api[_-]?key|authorization|cookie|password|secret|token)=([^&\s]+)", r"\1=[REDACTED]", text)
    text = re.sub(r"(?i)(bearer\s+)[a-z0-9._\-]+", r"\1[REDACTED]", text)
    return text


def _recommended_focus(progress: dict[str, Any], quality: dict[str, Any], supervision: dict[str, Any]) -> list[str]:
    focus: list[str] = []
    buckets = progress.get("failure_buckets") if isinstance(progress.get("failure_buckets"), dict) else {}
    if any(key in buckets for key in ("challenge_like", "captcha", "managed_challenge", "http_blocked")):
        focus.append("access_challenge")
    if int(progress.get("records_saved") or 0) == 0:
        focus.append("zero_records")
    if str(progress.get("quality_indicator") or "").lower() in {"fail", "unknown"}:
        focus.append("quality_repair")
    completeness = quality.get("field_completeness") if isinstance(quality.get("field_completeness"), dict) else {}
    if any(isinstance(value, (int, float)) and value < 0.8 for value in completeness.values()):
        focus.append("field_coverage")
    last = supervision.get("last_event") if isinstance(supervision.get("last_event"), dict) else {}
    if last.get("action"):
        focus.append(f"supervision_{last.get('action')}")
    return list(dict.fromkeys(focus))[:10]


def _compact_dict(value: Any, *, max_items: int) -> dict[str, Any]:
    if not isinstance(value, dict):
        return {}
    output: dict[str, Any] = {}
    for index, (key, item) in enumerate(value.items()):
        if index >= max_items:
            output["_truncated"] = True
            break
        if isinstance(item, dict):
            output[str(key)] = _compact_dict(item, max_items=max_items)
        elif isinstance(item, list):
            output[str(key)] = item[:10]
        else:
            output[str(key)] = item
    return output


def events_for_job(job: dict[str, Any]) -> list[dict[str, Any]]:
    events = [
        {
            "time": job.get("created_at", ""),
            "type": "job_created",
            "message": f"{job.get('user_goal', 'job')} created",
            "data": {"target": job.get("target_url", "")},
        },
        {
            "time": job.get("updated_at", ""),
            "type": f"job_{job.get('status', 'unknown')}",
            "message": f"job status: {job.get('status', 'unknown')}",
            "data": summarize_run_progress(job),
        },
    ]
    result = job.get("profile_run") or job.get("multi_profile_run")
    if isinstance(result, dict) and result.get("failures"):
        for failure in list(result.get("failures") or [])[:50]:
            events.append({
                "time": job.get("updated_at", ""),
                "type": "failure",
                "message": str(failure.get("error") or "crawl item failed"),
                "data": failure,
            })
    export = job.get("export")
    if isinstance(export, dict):
        events.append({
            "time": job.get("updated_at", ""),
            "type": "export_ready" if not export.get("error") else "export_failed",
            "message": str(export.get("error") or f"export ready: {export.get('output_path', '')}"),
            "data": export,
        })
    for decision in list(job.get("ai_decisions") or [])[:50]:
        if not isinstance(decision, dict):
            continue
        stage = str(decision.get("stage") or "ai_decision")
        summary = str(decision.get("reasoning_summary") or decision.get("error") or stage)
        events.append({
            "time": decision.get("created_at") or job.get("updated_at", ""),
            "type": f"ai_{stage}",
            "message": summary,
            "data": decision,
        })
    for trace in list(job.get("llm_traces") or [])[:100]:
        if not isinstance(trace, dict):
            continue
        stage = str(trace.get("stage") or "llm")
        status = str(trace.get("status") or "")
        duration = trace.get("duration_ms")
        events.append({
            "time": trace.get("created_at") or job.get("updated_at", ""),
            "type": f"llm_trace_{stage}",
            "message": f"{stage} {status} ({duration}ms)",
            "data": trace,
        })
    for record in list(job.get("managed_actions") or [])[:20]:
        if not isinstance(record, dict):
            continue
        result = record.get("result") if isinstance(record.get("result"), dict) else {}
        plan = result.get("plan") if isinstance(result.get("plan"), dict) else {}
        actions = plan.get("actions") if isinstance(plan.get("actions"), list) else []
        events.append({
            "time": record.get("created_at") or job.get("updated_at", ""),
            "type": "managed_actions_executed" if record.get("executed") else "managed_actions_planned",
            "message": f"managed actions: {len(actions)}",
            "data": record,
        })
    for step in list(job.get("managed_steps") or [])[:20]:
        if not isinstance(step, dict):
            continue
        action_record = step.get("action_record") if isinstance(step.get("action_record"), dict) else {}
        result = action_record.get("result") if isinstance(action_record.get("result"), dict) else {}
        plan = result.get("plan") if isinstance(result.get("plan"), dict) else {}
        actions = plan.get("actions") if isinstance(plan.get("actions"), list) else []
        child_run = step.get("child_run") if isinstance(step.get("child_run"), dict) else {}
        message = f"managed step {step.get('stage', 'unknown')}: {len(actions)} actions"
        if child_run.get("task_id"):
            message += f", child={child_run.get('task_id')}"
        events.append({
            "time": step.get("created_at") or job.get("updated_at", ""),
            "type": "managed_step_executed",
            "message": message,
            "data": step,
        })
    for probe in list(job.get("access_probe_history") or [])[:20]:
        if not isinstance(probe, dict):
            continue
        snapshot = probe.get("snapshot") if isinstance(probe.get("snapshot"), dict) else {}
        summary = snapshot.get("summary") if isinstance(snapshot.get("summary"), dict) else {}
        runtime = str(summary.get("recommended_runtime") or "")
        events.append({
            "time": probe.get("created_at") or job.get("updated_at", ""),
            "type": "access_probe_completed",
            "message": f"access probe completed{': ' + runtime if runtime else ''}",
            "data": probe,
        })
    for control in list(job.get("managed_control_loops") or [])[:20]:
        if not isinstance(control, dict):
            continue
        timeline = control.get("timeline") if isinstance(control.get("timeline"), list) else []
        child_run = control.get("child_run") if isinstance(control.get("child_run"), dict) else {}
        message = f"managed control loop: {len(timeline)} stages"
        if child_run.get("task_id"):
            message += f", child={child_run.get('task_id')}"
        events.append({
            "time": control.get("created_at") or job.get("updated_at", ""),
            "type": "managed_control_loop_completed",
            "message": message,
            "data": control,
        })
    supervision = job.get("supervision") if isinstance(job.get("supervision"), dict) else {}
    last_event = supervision.get("last_event") if isinstance(supervision.get("last_event"), dict) else {}
    if last_event:
        action = str(last_event.get("action") or "supervision")
        events.append({
            "time": job.get("updated_at", ""),
            "type": f"supervision_{action}",
            "message": str(last_event.get("reason") or action),
            "data": supervision,
        })
    auto_repair = job.get("managed_auto_repair") if isinstance(job.get("managed_auto_repair"), dict) else {}
    if auto_repair:
        child_id = str(auto_repair.get("child_task_id") or "")
        reason = str(auto_repair.get("reason") or "managed auto repair")
        events.append({
            "time": auto_repair.get("created_at") or job.get("updated_at", ""),
            "type": "managed_auto_repair_started" if auto_repair.get("attempted") else "managed_auto_repair_skipped",
            "message": f"{reason}: {child_id}" if child_id else reason,
            "data": auto_repair,
        })
    return events


def export_product_records(
    *,
    run_id: str,
    runtime_dir: str,
    export_spec: ExportSpec,
) -> dict[str, Any]:
    product_db = Path(runtime_dir) / "products.sqlite3" if runtime_dir else None
    store = ProductStore(product_db)
    records = store.list_records(run_id, limit=1_000_000)
    rows = [_record_to_export_row(record, export_spec.field_mapping) for record in records]
    fmt = export_spec.format.lower()
    suffix = _export_suffix(fmt)
    output = Path(export_spec.output_path or f"dev_logs/exports/{run_id}.{suffix}")
    if output.suffix.lower() != f".{suffix}":
        output = output.with_suffix(f".{suffix}")
    output.parent.mkdir(parents=True, exist_ok=True)
    if fmt == "json":
        output.write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    elif fmt == "csv":
        _write_csv(output, rows)
    elif fmt in {"xlsx", "xls"}:
        _write_xlsx(output, rows, template_path=export_spec.template_path, template=export_spec.template)
    elif fmt in {"sqlite", "db"}:
        if product_db is None or not product_db.exists():
            raise ValueError("runtime_dir with products.sqlite3 is required for sqlite export")
        shutil.copyfile(product_db, output)
    else:
        raise ValueError("format must be json, csv, xlsx, sqlite, or db")
    return {
        "schema_version": "export-result/v1",
        "run_id": run_id,
        "format": fmt,
        "output_path": str(output),
        "record_count": len(rows),
    }


def site_profile_from_analysis(
    *,
    url: str,
    recon: dict[str, Any],
    catalog_nodes: list[dict[str, Any]],
    field_candidates: list[FieldCandidate],
) -> SiteProfile:
    dom = recon.get("dom_structure") if isinstance(recon.get("dom_structure"), dict) else {}
    selectors: dict[str, Any] = {}
    if dom.get("product_selector"):
        selectors["item_container"] = dom.get("product_selector")
    for field in field_candidates:
        if field.selector:
            selectors[field.name] = field.selector
    seed_urls = [node["url"] for node in _flatten_catalog_nodes(catalog_nodes) if node.get("url")]
    if not seed_urls and url:
        seed_urls = [url]
    return SiteProfile.from_dict({
        "name": urlparse(url).netloc or "analyzed-site",
        "selectors": selectors,
        "target_fields": [field.name for field in field_candidates if field.selected],
        "pagination_hints": {"type": "dom_links"},
        "access_config": _analysis_access_config(recon),
        "crawl_preferences": {
            "seed_urls": seed_urls[:500],
            "seed_kind": "list",
            "catalog_tree": catalog_nodes,
        },
        "quality_expectations": {
            "required_fields": ["title"],
            "field_thresholds": {"title": 0.8},
        },
    })


def _analysis_access_config(recon: dict[str, Any]) -> dict[str, Any]:
    rendering = str(recon.get("rendering") or "").strip().lower()
    framework = str(recon.get("frontend_framework") or "").strip().lower()
    dom = recon.get("dom_structure") if isinstance(recon.get("dom_structure"), dict) else {}
    item_count = int(dom.get("item_count") or 0)
    if rendering in {"spa", "pwa", "client"} or framework in {"react", "vue", "angular"} or item_count == 0:
        return {
            "mode": "dynamic",
            "wait_until": "domcontentloaded",
            "browser_config": {
                "render_time_ms": 5000,
                "auto_accept_cookies": True,
                "capture_api": True,
            },
        }
    return {}


def _catalog_nodes_from_any(value: Any, *, path: list[str], source: str) -> list[CatalogNode]:
    nodes: list[CatalogNode] = []
    if isinstance(value, dict):
        for label, child in value.items():
            text = _clean_text(str(label))
            child_path = [*path, text] if text else list(path)
            if isinstance(child, str):
                nodes.append(_leaf_node(text, child, child_path, source))
            elif isinstance(child, (dict, list)):
                children = _catalog_nodes_from_any(child, path=child_path, source=source)
                nodes.append(CatalogNode(
                    id=_stable_id(child_path, ""),
                    label=text,
                    path=child_path,
                    level1=_level(child_path, 0),
                    level2=_level(child_path, 1),
                    level3=_level(child_path, 2),
                    children=children,
                    source=source,
                ))
    elif isinstance(value, list):
        for item in value:
            if isinstance(item, dict) and ("url" in item or "label" in item):
                label = _clean_text(str(item.get("label") or item.get("name") or item.get("title") or item.get("url") or ""))
                item_path = [str(part) for part in (item.get("path") or [*path, label]) if str(part).strip()]
                children = _catalog_nodes_from_any(item.get("children"), path=item_path, source=source) if item.get("children") else []
                nodes.append(CatalogNode(
                    id=str(item.get("id") or _stable_id(item_path, str(item.get("url") or ""))),
                    label=label,
                    url=str(item.get("url") or ""),
                    path=item_path,
                    level1=str(item.get("level1") or _level(item_path, 0)),
                    level2=str(item.get("level2") or _level(item_path, 1)),
                    level3=str(item.get("level3") or _level(item_path, 2)),
                    children=children,
                    source=(
                        f"graphql:{item.get('graphql_category_uid')}"
                        if item.get("graphql_category_uid")
                        else str(item.get("source") or source)
                    ),
                ))
            elif isinstance(item, (dict, list)):
                nodes.extend(_catalog_nodes_from_any(item, path=path, source=source))
    return nodes


def _leaf_node(label: str, url: str, path: list[str], source: str) -> CatalogNode:
    return CatalogNode(
        id=_stable_id(path, url),
        label=label,
        url=url,
        path=path,
        level1=_level(path, 0),
        level2=_level(path, 1),
        level3=_level(path, 2),
        source=source,
    )


def _flatten_catalog_nodes(nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    flat: list[dict[str, Any]] = []
    for node in nodes or []:
        if not isinstance(node, dict):
            continue
        if node.get("url"):
            flat.append(node)
        flat.extend(_flatten_catalog_nodes(list(node.get("children") or [])))
    return flat


def _record_to_export_row(record: ProductRecord, mapping: dict[str, str]) -> dict[str, Any]:
    path = [part for part in str(record.category or "").split(">") if part.strip()]
    raw = {
        "title": record.title,
        "highest_price": record.highest_price,
        "currency": record.currency,
        "colors": "; ".join(record.colors),
        "sizes": "; ".join(record.sizes),
        "description": record.description,
        "image_urls": "; ".join(record.image_urls),
        "canonical_url": record.canonical_url,
        "source_url": record.source_url,
        "category": record.category,
        "category_level_1": _level(path, 0),
        "category_level_2": _level(path, 1),
        "category_level_3": _level(path, 2),
    }
    if not mapping:
        return raw
    return {target: raw.get(source, "") for source, target in mapping.items()}


def _write_csv(output: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = sorted({key for row in rows for key in row.keys()})
    with output.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(
    output: Path,
    rows: list[dict[str, Any]],
    *,
    template_path: str = "",
    template: ExportTemplate | None = None,
) -> None:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("pandas is required for xlsx export") from exc

    if template and (template.field_to_column or template.columns or template.start_row > 1 or template.start_column > 1):
        _write_xlsx_with_template(output, rows, template)
    else:
        pd.DataFrame(rows).to_excel(output, index=False, sheet_name=(template.sheet_name if template else "Sheet1"))


def _write_xlsx_with_template(
    output: Path,
    rows: list[dict[str, Any]],
    template: ExportTemplate,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for template xlsx export") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = template.sheet_name

    # Determine column order: explicit columns list takes precedence,
    # then field_to_column keys (ordered), then all row keys.
    if template.columns:
        field_order = list(template.columns)
    elif template.field_to_column:
        field_order = list(template.field_to_column.keys())
    else:
        field_order = sorted({key for row in rows for key in row.keys()})

    # Write header row at (start_row, start_column)
    for col_idx, field_name in enumerate(field_order):
        header_label = template.field_to_column.get(field_name, field_name)
        ws.cell(row=template.start_row, column=template.start_column + col_idx, value=header_label)

    # Write data rows starting at start_row + 1
    for row_offset, row_data in enumerate(rows):
        for col_idx, field_name in enumerate(field_order):
            value = row_data.get(field_name, "")
            ws.cell(row=template.start_row + 1 + row_offset, column=template.start_column + col_idx, value=value)

    wb.save(str(output))


def _fields_from_text(text: str) -> list[str]:
    lowered = str(text or "").lower()
    found: list[str] = []
    for canonical, aliases in FIELD_ALIASES.items():
        if any(alias.lower() in lowered for alias in aliases):
            found.append(canonical)
    return found


def _normalize_field_name(name: str) -> str:
    text = name.strip().lower()
    for canonical, aliases in FIELD_ALIASES.items():
        if text == canonical or text in {alias.lower() for alias in aliases}:
            return canonical
    return re.sub(r"[^a-z0-9_]+", "_", text).strip("_")


def _selector_for_field(name: str, selectors: dict[str, Any]) -> str:
    if name in selectors:
        return str(selectors[name])
    aliases = FIELD_ALIASES.get(name, set())
    for key, value in selectors.items():
        if key in aliases or str(key).lower() in {alias.lower() for alias in aliases}:
            return str(value)
    if name == "highest_price" and selectors.get("price"):
        return str(selectors["price"])
    if name == "image_urls" and selectors.get("image"):
        return str(selectors["image"])
    return ""


def _field_label(name: str) -> str:
    labels = {
        "title": "商品标题",
        "highest_price": "最高价格",
        "colors": "颜色",
        "sizes": "尺码",
        "description": "商品描述",
        "image_urls": "商品图URL",
        "canonical_url": "商品详情URL",
        "category_level_1": "一级目录",
        "category_level_2": "二级目录",
        "category_level_3": "三级目录",
    }
    return labels.get(name, name)


def _looks_like_catalog_url(url: str, label: str) -> bool:
    parsed = urlparse(str(url or ""))
    lower_path = parsed.path.lower()
    lower_all = f"{parsed.netloc} {parsed.path} {label}".lower()
    tokens = set(_catalog_match_tokens(parsed.path, label))
    positive_tokens = {
        "category", "categories", "collection", "collections", "catalog", "shop", "sklep",
        "product", "products", "produkty", "women", "woman", "men", "kids", "children",
        "sale", "outlet", "new", "nowosci", "kobieta", "mezczyzna", "akcesoria",
        "rower", "rowery", "odziez", "helmet", "helmets", "shoe", "shoes", "bike", "bikes",
    }
    negative_tokens = {
        "login", "signin", "sign", "account", "cart", "checkout", "privacy", "terms",
        "contact", "kontakt", "blog", "wishlist", "compare", "newsletter", "unsubscribe",
        "forgot", "password", "create", "graphql", "rest", "payment", "payments", "saved",
        "service", "services", "shipping", "delivery", "method", "methods", "order", "orders",
        "returns", "return", "help", "faq", "customer", "support",
    }
    negative_phrases = (
        "/graphql", "/rest/", "payment-information", "saved-payments", "payment-method",
        "bike-service", "customer-service", "privacy-policy", "terms-and-conditions",
    )
    if tokens & negative_tokens or any(phrase in lower_all for phrase in negative_phrases):
        return False
    if tokens & positive_tokens:
        return True
    return any(
        phrase in lower_path
        for phrase in ("/product-category/", "/c/", "/cat/", "/category/", "/collections/")
    )


def _catalog_match_tokens(url_path: str, label: str) -> list[str]:
    raw = f"{url_path} {label}"
    return [
        token
        for token in re.split(r"[^a-z0-9ąćęłńóśźż]+", raw.lower())
        if token
    ]


def _url_without_fragment(url: str) -> str:
    return str(url).split("#", 1)[0]


def _stable_id(path: list[str], url: str) -> str:
    import hashlib

    raw = " > ".join(path) + "|" + str(url)
    return hashlib.sha1(raw.encode("utf-8", errors="ignore")).hexdigest()[:16]


def _level(path: list[str], index: int) -> str:
    return str(path[index]) if len(path) > index else ""


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _label_from_url(url: str) -> str:
    parsed = urlparse(str(url or ""))
    path = parsed.path or str(url or "")
    parts = [part for part in path.strip("/").split("/") if part and not part.startswith(":")]
    value = parts[-1] if parts else parsed.netloc
    value = re.sub(r"[-_]+", " ", value)
    value = re.sub(r"\.(html?|php|aspx?)$", "", value, flags=re.I)
    return _clean_catalog_label(value)


def _clean_catalog_label(label: str) -> str:
    text = _clean_text(label)
    text = re.sub(r"\s+", " ", text.replace("%20", " "))
    return text[:80]


def _empty_recon(url: str) -> dict[str, Any]:
    return {
        "url": url,
        "framework": "unknown",
        "rendering": "unknown",
        "anti_bot": False,
        "dom_structure": {"item_count": 0, "field_selectors": {}, "product_selector": ""},
    }


def _export_suffix(fmt: str) -> str:
    value = str(fmt or "xlsx").lower()
    return "sqlite3" if value in {"sqlite", "db"} else value
